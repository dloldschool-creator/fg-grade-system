"""Generate blank Excel templates for the Import from Excel page.

Built from `app.import_specs` rather than a hand-written column list, so
the template cannot drift from what the importer actually accepts. Re-run
it whenever sections or subjects change and the reference sheet updates
with them.

    .venv\\Scripts\\python.exe -m scripts.make_import_templates

Writes into `import-templates/` next to the project.

**The one trap worth knowing.** Excel treats a 12-digit LRN as a number,
which shows it as 1.07041E+11 and silently destroys a leading zero. The
LRN column here is pre-formatted as text so that cannot happen while the
sheet is being filled in — but if data is pasted in from elsewhere, check
the LRNs afterwards.
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.database import SessionLocal
from app.import_specs import SPECS
from app.models.academic_structure import Section
from app.models.subjects import Subject

OUTPUT_DIR = os.path.abspath(
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "import-templates")
)

HEADER_FILL = PatternFill("solid", start_color="FF1B4F9C", end_color="FF1B4F9C")
REQUIRED_FONT = Font(bold=True, color="FFFFFFFF")
OPTIONAL_FONT = Font(bold=True, color="FFD6E0F0", italic=True)

# Columns that must be text, not numbers.
TEXT_FIELDS = {"lrn"}

# Fixed choice lists, offered as dropdowns so a typo can't reach validation.
CHOICES = {
    "sex": ["MALE", "FEMALE"],
    "term": ["1", "2", "3"],
}

EXAMPLES = {
    "last_name": "DELA CRUZ",
    "first_name": "JUAN MIGUEL",
    "middle_name": "SANTOS",
    "extension_name": "JR",
    "sex": "MALE",
    "birthdate": "2009-01-15",
    "lrn": "107041140016",
    "section": "STEM - A",
    "subject": "General Mathematics",
    "term": "1",
    "grade": "90",
}

ROWS_TO_PREPARE = 400


def _data_sheet(workbook, spec):
    worksheet = workbook.active
    worksheet.title = "Data"

    for index, column in enumerate(spec.columns, start=1):
        cell = worksheet.cell(1, index, column.label)
        cell.fill = HEADER_FILL
        cell.font = REQUIRED_FONT if column.required else OPTIONAL_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = max(len(column.label) + 4, 16)

        # Pre-format the whole column, not just the filled rows — the
        # point is to be correct before anyone types in it.
        if column.field in TEXT_FIELDS:
            for row in range(2, ROWS_TO_PREPARE + 2):
                worksheet.cell(row, index).number_format = "@"

        if column.field in CHOICES:
            rule = DataValidation(
                type="list",
                formula1='"%s"' % ",".join(CHOICES[column.field]),
                allow_blank=not column.required,
                showDropDown=False,
            )
            rule.error = "Choose one of: %s" % ", ".join(CHOICES[column.field])
            rule.errorTitle = "Not a valid value"
            worksheet.add_data_validation(rule)
            rule.add(f"{letter}2:{letter}{ROWS_TO_PREPARE + 1}")

    worksheet.freeze_panes = "A2"
    return worksheet


def _instructions_sheet(workbook, spec, key):
    worksheet = workbook.create_sheet("Instructions")
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 30
    worksheet.column_dimensions["D"].width = 62

    def line(row, *values, bold=False):
        for offset, value in enumerate(values):
            cell = worksheet.cell(row, offset + 1, value)
            if bold:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=offset == 3)

    line(1, spec.label + " — import template", bold=True)
    line(2, spec.description)
    line(4, "Column", "Required", "Example", "Notes", bold=True)

    notes = {
        "lrn": "12 digits, stored as text so leading zeros survive. Already formatted "
               "as text on the Data sheet — if you paste from elsewhere, check it did "
               "not become 1.07041E+11.",
        "sex": "MALE or FEMALE. Choose from the dropdown.",
        "birthdate": "YYYY-MM-DD, or a real Excel date. Impossible dates are rejected.",
        "section": "Optional on the learner sheet — fill it in and the learner is "
                   "enrolled at the same time, leave it blank and they are created "
                   "only. Required on the grades sheet. Must match a section name "
                   "exactly; see the Reference sheet.",
        "subject": "Must match a subject name exactly, and must be offered to that "
                   "section in that term. See the Reference sheet.",
        "term": "1, 2 or 3.",
        "grade": "Whole number 0-100. Leave blank if not yet encoded — a blank stays "
                 "blank and never becomes 0.",
        "middle_name": "Optional.",
        "extension_name": "Optional. Jr, Sr, III and so on.",
    }

    row = 5
    for column in spec.columns:
        line(row, column.label, "yes" if column.required else "optional",
             EXAMPLES.get(column.field, ""), notes.get(column.field, ""))
        row += 1

    row += 1
    line(row, "How to use", bold=True); row += 1
    for text in [
        "1. Fill in the Data sheet, one row per record. Do not rename or reorder the columns.",
        "2. Leave a cell blank when you do not have the value. Never type 0 or N/A to fill a gap.",
        "3. Save as .xlsx and upload on the Import from Excel page.",
        "4. The system validates first and shows every problem before anything is written.",
        "5. Nothing is imported until you press Confirm.",
    ]:
        line(row, "", "", "", text); row += 1

    if key.value == "TERM_GRADES":
        row += 1
        line(row, "Note", bold=True); row += 1
        line(row, "", "", "", "Imported grades arrive as DRAFT and still go through the "
                              "normal submit and verify steps. Re-importing the same "
                              "learner, subject and term updates that grade rather than "
                              "adding a duplicate.")
    return worksheet


def _reference_sheet(workbook, session):
    worksheet = workbook.create_sheet("Reference")
    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["C"].width = 40

    sections = [s.name for s in session.query(Section).order_by(Section.name).all()]
    subjects = [s.official_name for s in session.query(Subject).order_by(Subject.official_name).all()]

    worksheet.cell(1, 1, "Sections").font = Font(bold=True)
    worksheet.cell(1, 3, "Subjects").font = Font(bold=True)
    if sections:
        for index, name in enumerate(sections, start=2):
            worksheet.cell(index, 1, name)
    else:
        worksheet.cell(2, 1, "(no sections created yet)").font = Font(italic=True)
    for index, name in enumerate(subjects, start=2):
        worksheet.cell(index, 3, name)

    worksheet.cell(len(subjects) + 3, 3,
                   "Names must match exactly. Regenerate this template if the "
                   "catalog changes.").font = Font(italic=True)
    return worksheet


def main() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    session = SessionLocal()
    try:
        for key, spec in SPECS.items():
            workbook = openpyxl.Workbook()
            _data_sheet(workbook, spec)
            _instructions_sheet(workbook, spec, key)
            _reference_sheet(workbook, session)
            name = "%s-import-template.xlsx" % spec.label.replace(" ", "-")
            path = os.path.join(OUTPUT_DIR, name)
            workbook.save(path)
            print("wrote %s  (%d columns)" % (path, len(spec.columns)))
    finally:
        session.close()
    print("\nFill in the Data sheet only. The Instructions and Reference sheets "
          "are guidance and are ignored by the importer.")


if __name__ == "__main__":
    main()
