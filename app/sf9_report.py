"""School Form 9 generation (§35) — the learner's Report Card / Learner's
Permanent Performance Report, filled into the official DepEd template with
`openpyxl` so the exact layout, merges, fonts and logos are preserved.

Shares `app/excel_template.py` with SF2: the template comes from the same
master automation workbook on OneDrive, so it carries the same
external-link formulas (109 of them) that must all be stripped before
anything is written.

**One template serves both grade levels.** §35 asks for separate Grade 11
and Grade 12 templates, but the supplied file is already grade-aware — its
own formulas branch on `SETUP!$B$13 = 11` to decide whether to draw the
combined-language hierarchy. Since we replace those formulas with values
anyway, the distinction becomes purely data-driven: a Grade 11 enrollment
has a combined learning area and a Grade 12 one doesn't, and
`app/report_card.py` already produces the right rows for either.

The learning-area rows come from `app/report_card.py` — the same builder
the Grade Summary screen uses — so the §16 combined-language rule can't
drift between what a teacher sees and what a parent receives.
"""

import os
from datetime import date

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties
from sqlalchemy.orm import Session

from app.attendance_engine import ActiveWindow
from app.attendance_service import (
    active_window_for,
    class_days_in_month,
    months_with_class_days,
    summarize_month,
)
from app.excel_template import (
    anchor_map,
    assert_no_external_links,
    strip_external_formulas,
    write,
    write_ref,
)
from app.models.academic_structure import GradeLevel, Section, Strand, Track
from app.models.enums import CompletionStatus
from app.models.grades import AnnualGradeSummary
from app.models.learners import Enrollment, Learner
from app.models.organization import School, SchoolYear
from app.models.rbac import User
from app.report_card import build_learning_area_rows

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "sf-templates",
    "SF9-template-with-sample-data.xlsx",
)
SHEET_NAME = "SF9"

# --- Template layout (1-based) --------------------------------------------
CELL_REGION = "A4"
CELL_DIVISION = "A5"
CELL_SCHOOL_NAME = "A7"
CELL_SCHOOL_YEAR = "H9"
CELL_NAME = "B10"
CELL_AGE = "I10"
CELL_SEX = "L10"
CELL_LRN = "B11"
CELL_GRADE = "I11"
CELL_SECTION = "L11"
CELL_TRACK = "B12"

# Learning Progress and Achievement. Each row is A:G (name), H/I/J (terms),
# K (final grade), L:M (remarks).
LEARNING_AREA_FIRST_ROW = 20
LEARNING_AREA_LAST_ROW = 31
COL_LEARNING_AREA = 1  # A
COL_TERM = {1: 8, 2: 9, 3: 10}  # H, I, J
COL_FINAL_GRADE = 11  # K
COL_REMARKS = 12  # L

ROW_GENERAL_AVERAGE = 32

# Attendance record: one column per month, June through April, then Total.
ROW_ATTENDANCE_MONTH = 3
ROW_CLASS_DAYS = 4
ROW_DAYS_PRESENT = 5
ROW_DAYS_ABSENT = 6
COL_ATTENDANCE_FIRST = 16  # P
COL_ATTENDANCE_TOTAL = 27  # AA
# The template's month headers, in column order — the school year runs
# June to April, which is why the list starts at 6 and wraps.
ATTENDANCE_MONTHS = [6, 7, 8, 9, 10, 11, 12, 1, 2, 3, 4]
MONTH_ABBREVIATIONS = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

# Teacher's comments — the label blocks are O9/O12/O15, the writable
# comment areas are the merged blocks beside them.
CELL_TERM_COMMENT = {1: "R9", 2: "R12", 3: "R15"}

# Certificate of Transfer block.
CELL_ADMITTED_TO_GRADE = "Q26"
CELL_ELIGIBLE_FOR_GRADE = "S27"
CELL_TRANSFER_ADVISER = "V29"   # label "Adviser" sits under it at V30
CELL_TRANSFER_SCHOOL_HEAD = "O31"  # label "School Head" sits under it at O32

# A Grade 12 completer moves on to college rather than to a Grade 13.
NEXT_LEVEL_AFTER_GRADE_12 = "COLLEGE"


# Print setup. The template ships with no <pageSetup> element at all, so
# without this the 27-column card defaults to portrait at 100% and splits
# across pages.
PRINT_AREA = "A1:AA42"

# Column N drives the template's OWN conditional formatting, which
# blocks out the terms a subject isn't offered in. Its three digits are
# per-term flags — 111 = all three terms, 100 = Term 1 only, 10 = Term 2
# only, 1 = Term 3 only — and the rules read them back as:
#
#   H (Term 1): AND($A20<>"", INT($N20/100)=0)
#   I (Term 2): AND($A20<>"", MOD(INT($N20/10),10)=0)
#   J (Term 3): AND($A20<>"", MOD($N20,10)=0)
#
# So writing the right flags is all that's needed; the shading, its exact
# grey and the white text on top are the official template's own styling.
# Blanking this column (as a stray "print helper") makes every digit 0 and
# greys out every grade on the card — which is exactly what happened.
COL_TERM_OFFERED_FLAGS = 14  # N
TERM_FLAG_PLACE = {1: 100, 2: 10, 3: 1}

MAX_LEARNING_AREAS = LEARNING_AREA_LAST_ROW - LEARNING_AREA_FIRST_ROW + 1


def _age_on(birthdate: date | None, reference: date | None) -> int | None:
    """Age as of the school year's start, not today — a report card
    reprinted years later must still show the age the learner was."""
    if birthdate is None or reference is None:
        return None
    years = reference.year - birthdate.year
    if (reference.month, reference.day) < (birthdate.month, birthdate.day):
        years -= 1
    return years


def _full_name(learner: Learner) -> str:
    middle = f" {learner.middle_name}" if learner.middle_name else ""
    extension = f" {learner.extension_name}" if learner.extension_name else ""
    return f"{learner.last_name}, {learner.first_name}{middle}{extension}".strip().upper()


def _monthly_attendance(session: Session, enrollment: Enrollment, window: ActiveWindow) -> dict:
    """(month) -> (class_days, present, absent) for months that actually
    have attendance encoded.

    Two rules worth keeping:

    - Class days are the learner's **eligible** days, not the section's
      calendar total — a late enrollee has none before their effective
      date (§31). Using the section total would show a learner absent for
      weeks before they even arrived.
    - A month with nothing encoded is **omitted entirely**, not reported
      as zero days present. §35 says to populate this from the attendance
      data; a month nobody has encoded has no data, and printing "22
      class days, 0 present" on a report card going home to a parent
      would read as the learner having missed the whole month.
    """
    result: dict[int, tuple[int, int, int]] = {}
    for year, month in months_with_class_days(session, enrollment.school_year_id):
        class_days = class_days_in_month(session, enrollment.school_year_id, year, month)
        summary = summarize_month(session, enrollment, window, class_days)
        if summary.eligible_days == 0 or summary.unencoded_days == summary.eligible_days:
            continue
        result[month] = (summary.eligible_days, summary.days_present, summary.days_absent)
    return result


def build_sf9_workbook(session: Session, enrollment_id):
    """Returns a self-contained openpyxl Workbook for one learner."""
    enrollment = session.get(Enrollment, enrollment_id)
    learner = session.get(Learner, enrollment.learner_id)
    section = session.get(Section, enrollment.section_id)
    school = session.query(School).one_or_none()
    school_year = session.get(SchoolYear, enrollment.school_year_id)
    grade_level = (
        session.get(GradeLevel, enrollment.grade_level_id) if enrollment.grade_level_id else None
    )
    track = session.get(Track, section.track_id) if section and section.track_id else None
    strand = session.get(Strand, section.strand_id) if section and section.strand_id else None

    rows = build_learning_area_rows(session, enrollment)
    if len(rows) > MAX_LEARNING_AREAS:
        raise ValueError(
            f"{_full_name(learner)} has {len(rows)} learning-area rows but the SF9 "
            f"template provides only {MAX_LEARNING_AREAS}."
        )

    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    strip_external_formulas(worksheet)
    anchors = anchor_map(worksheet)

    _fill_identity(
        worksheet, anchors,
        school=school, school_year=school_year, learner=learner, section=section,
        grade_level=grade_level, track=track, strand=strand,
    )
    _fill_learning_areas(worksheet, anchors, rows)
    _fill_general_average(session, worksheet, anchors, enrollment)
    _fill_attendance(session, worksheet, anchors, enrollment)
    _fill_transfer_certificate(session, worksheet, anchors, enrollment, section, grade_level)
    _fill_comments(worksheet, anchors, enrollment)
    _apply_print_setup(worksheet)

    workbook._external_links = []
    assert_no_external_links(workbook)
    return workbook


def _fill_identity(worksheet, anchors, *, school, school_year, learner, section,
                   grade_level, track, strand) -> None:
    if school is not None:
        write_ref(worksheet, anchors, CELL_REGION, school.region)
        write_ref(worksheet, anchors, CELL_DIVISION, school.schools_division)
        write_ref(worksheet, anchors, CELL_SCHOOL_NAME, school.school_name)
    write_ref(worksheet, anchors, CELL_SCHOOL_YEAR, school_year.name if school_year else "")
    write_ref(worksheet, anchors, CELL_NAME, _full_name(learner))
    write_ref(
        worksheet, anchors, CELL_AGE,
        _age_on(learner.birthdate, school_year.start_date if school_year else None),
    )
    write_ref(worksheet, anchors, CELL_SEX, learner.sex.value.title())
    # LRN is text, never numeric — leading zeros must survive (rule 10).
    write_ref(worksheet, anchors, CELL_LRN, learner.lrn or "")
    write_ref(worksheet, anchors, CELL_GRADE, _grade_level_number(grade_level))
    write_ref(worksheet, anchors, CELL_SECTION, section.name if section else "")
    write_ref(
        worksheet, anchors, CELL_TRACK,
        " / ".join(p for p in [track.name if track else "", strand.name if strand else ""] if p),
    )
    # "Academic / Science, Technology, Engineering, and Mathematics" is
    # longer than the B12:J12 merge, and a merged cell can't overflow into
    # its neighbours the way a plain one does — so it was being cut off.
    # shrink_to_fit makes Excel scale the font down instead of truncating,
    # which keeps the full strand name readable without altering the
    # template's column widths.
    _shrink_to_fit(worksheet, anchors, CELL_TRACK)


def _centre(worksheet, anchors, coordinate: str) -> None:
    cell = worksheet[coordinate]
    target = anchors.get((cell.row, cell.column), (cell.row, cell.column))
    target_cell = worksheet.cell(*target)
    existing = target_cell.alignment
    target_cell.alignment = Alignment(
        horizontal="center",
        vertical=existing.vertical,
        wrap_text=existing.wrap_text,
        shrink_to_fit=existing.shrink_to_fit,
    )


def _apply_print_setup(worksheet) -> None:
    """Landscape, scaled so the whole card lands on a single sheet.

    `fitToWidth`/`fitToHeight` only take effect when the sheet's
    `fitToPage` property is also set — setting the page-setup fields
    alone silently does nothing.

    The margins are narrowed first, and that matters more than it looks:
    the card's content is about 9.5 inches tall, while landscape Letter
    inside the template's original 0.75-inch margins leaves only 7.0
    inches — forcing Excel to shrink everything to roughly 73% to fit one
    page. Trimming the margins to 0.25 inch gives about 8.0 inches of
    usable height, so the same content fits at roughly 84% and stays
    comfortably readable.

    Paper size is left to the printer's default, since the template
    doesn't specify one and DepEd offices print on A4, Letter or Folio.
    """
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1  # one sheet, not "as many as needed"
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.print_area = PRINT_AREA

    margins = worksheet.page_margins
    margins.left = margins.right = 0.25
    margins.top = margins.bottom = 0.25
    margins.header = margins.footer = 0.15

    # Centre the card on the sheet. Fitting to one page scales by
    # whichever dimension binds first — here it's the height, since the
    # card is proportionally taller than the page — which leaves slack
    # across the width and parks everything against the left margin.
    # These two flags distribute that leftover space evenly instead.
    worksheet.print_options.horizontalCentered = True
    worksheet.print_options.verticalCentered = True


def _shrink_to_fit(worksheet, anchors, coordinate: str) -> None:
    cell = worksheet[coordinate]
    target = anchors.get((cell.row, cell.column), (cell.row, cell.column))
    target_cell = worksheet.cell(*target)
    existing = target_cell.alignment
    target_cell.alignment = Alignment(
        horizontal=existing.horizontal,
        vertical=existing.vertical,
        wrap_text=False,
        shrink_to_fit=True,
    )


def _grade_level_number(grade_level) -> str:
    """The form's field is labelled "Grade:", so it wants "11" — the
    stored code is "G11"."""
    if grade_level is None:
        return ""
    digits = "".join(ch for ch in (grade_level.code or grade_level.name or "") if ch.isdigit())
    return digits or (grade_level.code or "")


def _fill_learning_areas(worksheet, anchors, rows) -> None:
    """Writes each row, then blanks the unused ones.

    §35: "Do not print unused placeholder subjects. Blank unused rows may
    remain visually blank to preserve the official template" — so the
    leftover rows are cleared rather than hidden, keeping the form's
    printed shape intact.
    """
    for index in range(MAX_LEARNING_AREAS):
        row_number = LEARNING_AREA_FIRST_ROW + index
        if index < len(rows):
            entry = rows[index]
            write(worksheet, anchors, row_number, COL_LEARNING_AREA, entry.display_name)
            for term_number, column in COL_TERM.items():
                offered = entry.is_offered(term_number)
                write(
                    worksheet, anchors, row_number, column,
                    _grade(entry.term_grades.get(term_number)) if offered else None,
                )
            # The template's conditional formatting blocks out whichever
            # terms these flags mark as not offered.
            write(worksheet, anchors, row_number, COL_TERM_OFFERED_FLAGS, _term_flags(entry))
            # A component's Final Grade and Remark stay blank — §16.
            write(worksheet, anchors, row_number, COL_FINAL_GRADE, _grade(entry.final_grade))
            write(worksheet, anchors, row_number, COL_REMARKS, entry.remark)
        else:
            write(worksheet, anchors, row_number, COL_LEARNING_AREA, None)
            for column in COL_TERM.values():
                write(worksheet, anchors, row_number, column, None)
            # An unused row stays visually blank (§35). The rules are all
            # guarded by $A<>"", so an empty name row never shades.
            write(worksheet, anchors, row_number, COL_TERM_OFFERED_FLAGS, None)
            write(worksheet, anchors, row_number, COL_FINAL_GRADE, None)
            write(worksheet, anchors, row_number, COL_REMARKS, None)


def _grade(value):
    """Grades are whole numbers by construction (§60) — printed as ints so
    the form doesn't show "93.00"."""
    return int(value) if value is not None else None


def _term_flags(entry) -> int:
    """The template's per-term offered flags for one row, as the 3-digit
    number its conditional formatting decodes (111 = all three terms
    offered, 100 = Term 1 only, and so on)."""
    return sum(place for term, place in TERM_FLAG_PLACE.items() if entry.is_offered(term))


def _fill_general_average(session, worksheet, anchors, enrollment) -> None:
    summary = (
        session.query(AnnualGradeSummary).filter_by(enrollment_id=enrollment.id).one_or_none()
    )
    write(worksheet, anchors, ROW_GENERAL_AVERAGE, COL_FINAL_GRADE,
          _grade(summary.general_average) if summary else None)
    remark = None
    if summary and summary.general_average is not None:
        remark = "PASSED" if summary.lowest_final_grade is not None and summary.failed_subject_count == 0 else "FAILED"
    write(worksheet, anchors, ROW_GENERAL_AVERAGE, COL_REMARKS, remark)


def _fill_attendance(session, worksheet, anchors, enrollment) -> None:
    """§35: populate from the attendance data, never re-encoded by hand."""
    window = active_window_for(session, enrollment)
    monthly = _monthly_attendance(session, enrollment, window)

    totals = [0, 0, 0]
    for index, month in enumerate(ATTENDANCE_MONTHS):
        column = COL_ATTENDANCE_FIRST + index
        # The month headers are external formulas in the template, so
        # they're stripped along with everything else and have to be
        # written back.
        write(worksheet, anchors, ROW_ATTENDANCE_MONTH, column, MONTH_ABBREVIATIONS[month])

        figures = monthly.get(month)
        if figures is None:
            for row in (ROW_CLASS_DAYS, ROW_DAYS_PRESENT, ROW_DAYS_ABSENT):
                write(worksheet, anchors, row, column, None)
            continue

        class_days, present, absent = figures
        write(worksheet, anchors, ROW_CLASS_DAYS, column, class_days)
        write(worksheet, anchors, ROW_DAYS_PRESENT, column, present)
        write(worksheet, anchors, ROW_DAYS_ABSENT, column, absent)
        totals[0] += class_days
        totals[1] += present
        totals[2] += absent

    # Totals cover only the months actually shown, so they always add up
    # to what's printed above them.
    for row, total in zip((ROW_CLASS_DAYS, ROW_DAYS_PRESENT, ROW_DAYS_ABSENT), totals):
        write(worksheet, anchors, row, COL_ATTENDANCE_TOTAL, total or None)


def _passed_everything(summary) -> bool:
    """Completed the year with nothing below the passing grade."""
    return (
        summary is not None
        and summary.completion_status == CompletionStatus.COMPLETE
        and (summary.failed_subject_count or 0) == 0
    )


def next_level_after(grade_number: str) -> str:
    """What a completer is eligible to enter next. Grade 12 is the end of
    Senior High, so its completers move on to COLLEGE rather than a
    non-existent Grade 13."""
    if not grade_number.isdigit():
        return ""
    number = int(grade_number)
    return NEXT_LEVEL_AFTER_GRADE_12 if number >= 12 else str(number + 1)


def _fill_transfer_certificate(session, worksheet, anchors, enrollment, section, grade_level) -> None:
    """The Certificate of Transfer block.

    "Admitted to Grade" is a fact of this year's enrolment, so it's always
    filled. "Eligible for Admission to Grade" is a judgement about the
    learner's result, so it's only filled once they've actually completed
    the year with no failed subject — leaving it blank otherwise rather
    than asserting an eligibility they haven't earned.
    """
    summary = (
        session.query(AnnualGradeSummary).filter_by(enrollment_id=enrollment.id).one_or_none()
    )
    grade_number = _grade_level_number(grade_level)

    write_ref(worksheet, anchors, CELL_ADMITTED_TO_GRADE, grade_number or None)
    write_ref(
        worksheet, anchors, CELL_ELIGIBLE_FOR_GRADE,
        next_level_after(grade_number) or None if _passed_everything(summary) else None,
    )

    school = session.query(School).one_or_none()
    adviser = (
        session.get(User, section.adviser_user_id)
        if section is not None and section.adviser_user_id
        else None
    )
    write_ref(
        worksheet, anchors, CELL_TRANSFER_ADVISER,
        adviser.full_name.upper() if adviser else None,
    )
    # The template centres the "Adviser" label beneath it and both
    # school-head cells, but leaves this one unaligned.
    _centre(worksheet, anchors, CELL_TRANSFER_ADVISER)
    write_ref(
        worksheet, anchors, CELL_TRANSFER_SCHOOL_HEAD,
        (school.school_head_name or "").upper() if school else None,
    )


def _fill_comments(worksheet, anchors, enrollment) -> None:
    comments = {
        1: enrollment.term1_adviser_comment,
        2: enrollment.term2_adviser_comment,
        3: enrollment.term3_adviser_comment,
    }
    for term_number, coordinate in CELL_TERM_COMMENT.items():
        write_ref(worksheet, anchors, coordinate, comments.get(term_number) or None)
