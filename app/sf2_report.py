"""School Form 2 generation (§34) — fills the official DepEd SF2 template
with `openpyxl` so the exact cell layout, merges, fonts and logos are
preserved, per the stack decision in CLAUDE.md.

**The template arrives tethered to another workbook.** `sf-templates/
SF2-template-with-sample-data.xlsx` is a print-view sheet whose ~1600
data cells are external-link formulas pointing at the school's master
automation workbook on OneDrive (`'[1]ATTENDANCE DAILY'!...`,
`'[1]SETUP'!...`). Every one of those is stripped before anything is
written, and the external-link definition itself is dropped, so a
generated file is fully self-contained: no "update links?" prompt, no
stale cached values, and no dependency on a file the app can't see.
`assert_no_external_links()` is the guard that keeps it that way.

Everything the form displays is computed by us (`app/attendance_engine.py`
via `app/attendance_service.py`), never read back from the template's own
formulas — the same "computation is ours, server-side" rule the grading
side follows.
"""

import calendar as _calendar
import io
import os
from datetime import date, timedelta

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from sqlalchemy.orm import Session

from app.excel_template import (
    anchor_map,
    assert_no_external_links,
    clear_column,
    replicate_images,
    strip_external_formulas,
    workbook_to_bytes,
    write,
    write_ref,
)
from app.attendance_service import (
    class_days_in_month,
    movements_by_enrollment,
    roster_for_month,
    summarize_month_batch,
)
from app.models.academic_structure import GradeLevel, Section, Strand, Track
from app.models.attendance import AttendanceRecord
from app.models.enums import AttendanceStatus, EnrollmentStatus, Sex
from app.models.organization import School, SchoolYear, Term
from app.models.rbac import User

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "sf-templates",
    "SF2-template-with-sample-data.xlsx",
)
SHEET_NAME = "SF2"

# --- Template layout ------------------------------------------------------
# Column indices are 1-based. The day columns are NOT contiguous: each day
# is a merged block of varying width, so the anchors have to be listed
# explicitly rather than computed from a start offset.
COL_NO = 1  # A
COL_NAME = 7  # G
DAY_COLS = [11, 13, 14, 15, 19, 22, 23, 26, 27, 28, 31, 34, 35, 38, 40,
            42, 44, 45, 47, 48, 49, 51, 54, 56, 57]
COL_ABSENT = 61  # BI
COL_PRESENT = 64  # BL
COL_REMARKS = 68  # BP

# Helper column the source workbook used to drive its own print macros —
# it holds "Select Month in" and a HYPERLINK to a 'PRINT CONTROL' sheet
# that doesn't exist here. The hyperlink is a *local* formula, so unlike
# the [1]-prefixed ones it survives external-link stripping and has to be
# cleared explicitly.
COL_PRINT_CONTROL = 81  # CC

# Rightmost column carrying real form content (the summary box's merges
# run out to CA). Everything past this is helper scaffolding, so the print
# area stops here.
COL_LAST_PRINTED = 79  # CA
ROW_LAST_PRINTED = 112

# Scaffolding row: in the source workbook this held date serials that the
# day-number and weekday rows read via formula (`=DAY(K$14)` and
# `=CHOOSE(WEEKDAY(K$14,2),…)`). We write those two rows as plain values
# instead, so nothing references row 14 any more — it's cleared rather
# than populated, otherwise raw dates print above the table header.
ROW_DATE_SCAFFOLD = 14
ROW_DAY_NUMBER = 16
ROW_WEEKDAY = 17
MALE_FIRST_ROW, MALE_LAST_ROW = 18, 42
ROW_MALE_TOTAL = 43
FEMALE_FIRST_ROW, FEMALE_LAST_ROW = 44, 68
ROW_FEMALE_TOTAL = 69
ROW_COMBINED_TOTAL = 70

ROWS_PER_SEX = MALE_LAST_ROW - MALE_FIRST_ROW + 1  # 25, per §34

# Summary box: M / F / TOTAL columns, and the row each figure sits on.
COL_SUMMARY_M, COL_SUMMARY_F, COL_SUMMARY_TOTAL = 73, 75, 76  # BU / BW / BX
ROW_ENROLMENT_FIRST_FRIDAY = 74
ROW_LATE_ENROLMENT = 77
ROW_REGISTERED_END = 82
ROW_PCT_ENROLMENT = 84
ROW_AVG_DAILY_ATTENDANCE = 86
ROW_PCT_ATTENDANCE = 88
ROW_FIVE_CONSECUTIVE = 89
ROW_NLS = 92
ROW_TRANSFERRED_OUT = 95
ROW_TRANSFERRED_IN = 98
ROW_SHIFTED_OUT = 100
ROW_SHIFTED_IN = 101

CELL_SCHOOL_NAME = "I5"
CELL_SCHOOL_ID = "Y5"
CELL_DISTRICT = "AQ5"
CELL_DIVISION = "BH5"
CELL_REGION = "BV5"
CELL_TERM = "I7"
CELL_TRACK_STRAND = "BC7"
CELL_SCHOOL_YEAR = "U8"
CELL_GRADE_LEVEL = "AN8"
CELL_MONTH = "BN11"
CELL_SECTION = "I12"
CELL_DAYS_OF_CLASSES = "BO71"
CELL_ADVISER = "BJ107"
CELL_SCHOOL_HEAD = "BK111"

WEEKDAY_LETTERS = {0: "M", 1: "T", 2: "W", 3: "TH", 4: "F", 5: "SA", 6: "SU"}

# §30's printed codes. PRESENT prints as a BLANK cell — that's the paper
# form's own convention ("(blank) - Present; X - Absent"), and the reason
# the encoding UI uses an explicit "P" instead: on screen we must be able
# to tell "encoded present" from "nobody has said yet", but on the printed
# form present is simply left empty.
PRINTED_CODES = {
    AttendanceStatus.PRESENT: "",
    AttendanceStatus.ABSENT: "X",
    AttendanceStatus.LATE: "T-L",
    AttendanceStatus.CUTTING: "T-C",
}

MOVEMENT_REMARKS = {
    EnrollmentStatus.TRANSFERRED_OUT: "Transferred Out",
    EnrollmentStatus.TRANSFERRED_IN: "Transferred In",
    EnrollmentStatus.NLS: "NLS",
    EnrollmentStatus.DROPPED: "Dropped",
    EnrollmentStatus.SHIFTED_OUT: "Shifted Out",
    EnrollmentStatus.SHIFTED_IN: "Shifted In",
    EnrollmentStatus.LATE_ENROLLMENT: "Late Enrollment",
}


# --- Pure helpers (unit-tested; no DB, no workbook) -----------------------


def paginate(male_count: int, female_count: int) -> int:
    """How many SF2 pages this section needs.

    §34: the form provides 25 male and 25 female rows, and overflow must
    produce additional pages rather than hiding learners. Male and female
    each get their own 25-row block on every page, so the page count is
    driven by whichever sex overflows further — not by the combined
    total.
    """
    pages_for_male = max(1, -(-male_count // ROWS_PER_SEX))  # ceil
    pages_for_female = max(1, -(-female_count // ROWS_PER_SEX))
    return max(pages_for_male, pages_for_female)


def page_slice(learners: list, page_index: int) -> list:
    """The learners belonging on `page_index` (0-based) for one sex."""
    start = page_index * ROWS_PER_SEX
    return learners[start : start + ROWS_PER_SEX]


def printed_code(status: AttendanceStatus | None) -> str:
    """Blank for present *and* for an un-encoded day — on paper they look
    the same. The distinction is enforced before printing instead: §33's
    pre-finalization check refuses to finalize a month with un-encoded
    days, so a finalized SF2 has no silent gaps."""
    if status is None:
        return ""
    return PRINTED_CODES.get(status, "")


def movement_remark(movement_type: EnrollmentStatus, effective: date) -> str:
    label = MOVEMENT_REMARKS.get(movement_type, movement_type.value.replace("_", " ").title())
    return f"{label} {effective:%m/%d/%Y}"


def first_friday_on_or_after(start: date) -> date:
    """The reference date for the form's "Enrolment as of (1st Friday of
    June)" line.

    Anchored to the school year's own start date rather than to June 1,
    because the calendar first Friday can fall *before* classes begin —
    SY 2026-2027 opens Monday 8 June, so the calendar first Friday (the
    5th) is a day on which nobody is enrolled yet and would report an
    enrolment of zero. The first Friday on or after opening (the 12th) is
    what the figure actually means.
    """
    return start + timedelta(days=(4 - start.weekday()) % 7)  # Friday == 4


# --- Template handling ----------------------------------------------------


def _learner_rows(session: Session, section_id, school_year_id, year: int, month: int, class_days):
    """One dict per learner on this month's sheet, already split by sex and
    in SF2's order (male then female, alphabetical) — `roster_for_month`
    handles that ordering and the §32 visibility rule."""
    roster = roster_for_month(session, section_id, school_year_id, year, month)
    day_ids = [d.id for d in class_days]
    records = {}
    if roster and day_ids:
        rows = (
            session.query(AttendanceRecord)
            .filter(
                AttendanceRecord.enrollment_id.in_([e.id for e, _, _ in roster]),
                AttendanceRecord.calendar_date_id.in_(day_ids),
            )
            .all()
        )
        records = {(r.enrollment_id, r.calendar_date_id): r.status for r in rows}

    month_start = date(year, month, 1)
    month_end = date(year, month, _calendar.monthrange(year, month)[1])

    summaries = summarize_month_batch(session, roster, class_days)
    movements = movements_by_enrollment(session, [e.id for e, _, _ in roster])

    males, females = [], []
    for enrollment, learner, window in roster:
        marks = []
        for day in class_days:
            if not window.contains(day.calendar_date):
                marks.append("")  # outside their enrolment window
                continue
            marks.append(printed_code(records.get((enrollment.id, day.id))))

        summary = summaries[enrollment.id]

        remarks = [
            movement_remark(m.movement_type, m.effective_date)
            for m in sorted(movements.get(enrollment.id, []), key=lambda m: m.effective_date)
            if month_start <= m.effective_date <= month_end
        ]

        row = {
            "name": _display_name(learner),
            "sex": learner.sex,
            "marks": marks,
            "absent": summary.days_absent,
            "present": summary.days_present,
            "remarks": "; ".join(remarks),
            "window": window,
            "enrollment": enrollment,
            "summary": summary,
        }
        (males if learner.sex == Sex.MALE else females).append(row)
    return males, females


def _display_name(learner) -> str:
    """Uppercased at render time as well as on save. Names are normalized
    to uppercase when stored (`normalize_name`), but this keeps any row
    predating that rule — or edited directly in the database — printing
    consistently with the rest of the form."""
    middle = f" {learner.middle_name}" if getattr(learner, "middle_name", None) else ""
    extension = f" {learner.extension_name}" if getattr(learner, "extension_name", None) else ""
    return f"{learner.last_name}, {learner.first_name}{middle}{extension}".strip().upper()


def _movement_counts(session: Session, rows, year: int, month: int) -> dict:
    """Counts of each movement type effective within the month, split by
    sex — the SF2 summary box's movement lines."""
    month_start = date(year, month, 1)
    month_end = date(year, month, _calendar.monthrange(year, month)[1])
    movements = movements_by_enrollment(session, [row["enrollment"].id for row in rows])
    counts: dict = {}
    for row in rows:
        for movement in movements.get(row["enrollment"].id, []):
            if not (month_start <= movement.effective_date <= month_end):
                continue
            bucket = counts.setdefault(movement.movement_type, {"M": 0, "F": 0})
            bucket["M" if row["sex"] == Sex.MALE else "F"] += 1
    return counts


# --- Workbook building ----------------------------------------------------


def build_sf2_workbook(
    session: Session, section_id, school_year_id, year: int, month: int
):
    """Returns a self-contained openpyxl Workbook for one section-month."""
    school = session.query(School).one_or_none()
    section = session.get(Section, section_id)
    school_year = session.get(SchoolYear, school_year_id)
    class_days = class_days_in_month(session, school_year_id, year, month)
    males, females = _learner_rows(session, section_id, school_year_id, year, month, class_days)

    if len(class_days) > len(DAY_COLS):
        raise ValueError(
            f"{_calendar.month_name[month]} {year} has {len(class_days)} class days but the "
            f"SF2 template provides only {len(DAY_COLS)} day columns."
        )

    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    base = workbook[SHEET_NAME]
    page_count = paginate(len(males), len(females))

    sheets = [base]
    for index in range(1, page_count):
        extra = workbook.copy_worksheet(base)
        extra.title = f"{SHEET_NAME} p{index + 1}"
        sheets.append(extra)
    if page_count > 1:
        base.title = f"{SHEET_NAME} p1"
        replicate_images(base, sheets[1:])

    term = _term_for_month(session, school_year_id, class_days)
    adviser = session.get(User, section.adviser_user_id) if section.adviser_user_id else None
    track, strand = _track_and_strand(session, section)

    for page_index, worksheet in enumerate(sheets):
        strip_external_formulas(worksheet)
        anchors = anchor_map(worksheet)
        _fill_header(
            session,
            worksheet,
            anchors,
            school=school,
            section=section,
            school_year=school_year,
            term=term,
            track=track,
            strand=strand,
            year=year,
            month=month,
            class_day_count=len(class_days),
            adviser=adviser,
        )
        _fill_day_headers(worksheet, anchors, class_days)
        _fill_learner_block(worksheet, anchors, page_slice(males, page_index), MALE_FIRST_ROW, class_days)
        _fill_learner_block(worksheet, anchors, page_slice(females, page_index), FEMALE_FIRST_ROW, class_days)
        _fill_daily_totals(worksheet, anchors, males, females, class_days)
        _fill_summary(session, worksheet, anchors, males, females, class_days, school_year, year, month)
        _clear_print_control_column(worksheet)
        _widen_summary_percentage_columns(worksheet)
        _apply_print_setup(worksheet)

    # Drop the link to the school's master workbook entirely, so the file
    # opens without an "update links?" prompt.
    workbook._external_links = []
    assert_no_external_links(workbook)
    return workbook


def _clear_print_control_column(worksheet) -> None:
    """Blanks the source workbook's print-macro helper column — it holds
    "Select Month in" and a HYPERLINK to a 'PRINT CONTROL' sheet that
    doesn't exist here. That hyperlink is a *local* formula, so unlike
    the [1]-prefixed ones it survives external-link stripping."""
    clear_column(worksheet, COL_PRINT_CONTROL)


def _widen_summary_percentage_columns(worksheet) -> None:
    """Widens the summary box's M and F sub-columns so percentages fit.

    Excel shows `###` when a *numeric* cell is too narrow (unlike text,
    which just overflows into the neighbour). The template gives the M
    figure BU+BV = 3.9 characters and the F figure BW = 3.8, which is
    fine for the counts but not for "100.00%" at seven characters — so
    both percentage rows rendered as ###. The Total column (BX:CA, ~15.6)
    was always wide enough, which is why only two of the three showed it.

    Widening rather than dropping to a `0%` format keeps the two decimals
    the official form reports for Percentage of Attendance. The form
    prints fit-to-width, so the extra ~7 characters just scale the sheet
    very slightly.
    """
    worksheet.column_dimensions[get_column_letter(COL_SUMMARY_M + 1)].width = 7.5  # BV
    worksheet.column_dimensions[get_column_letter(COL_SUMMARY_F)].width = 7.5  # BW


def _apply_print_setup(worksheet) -> None:
    """Landscape, scaled to fit the form's width on one page.

    The template ships with **no `<pageSetup>` element at all**, so
    without this Excel falls back to portrait at 100% and the 79-column
    form spills across several pages. `fitToWidth` only takes effect when
    the sheet's `fitToPage` property is also set — setting the page-setup
    field alone silently does nothing.

    Height is left unconstrained (`fitToHeight = 0`) so a long form can
    still run onto a second sheet of paper rather than being squashed
    illegibly.

    Paper size is deliberately not set: the template doesn't specify one,
    and DepEd forms get printed on A4, Letter or Folio depending on the
    office, so whatever the printer defaults to is a better guess than
    anything hardcoded here.
    """
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.print_area = (
        f"A1:{get_column_letter(COL_LAST_PRINTED)}{ROW_LAST_PRINTED}"
    )


def _term_for_month(session: Session, school_year_id, class_days):
    """The term this month's class days belong to. A month can straddle
    two terms (§29's September split), in which case the term covering the
    most class days wins for the header field."""
    if not class_days:
        return None
    tally: dict = {}
    for day in class_days:
        if day.term_id is not None:
            tally[day.term_id] = tally.get(day.term_id, 0) + 1
    if not tally:
        return None
    term_id = max(tally, key=tally.get)
    return session.get(Term, term_id)


def _track_and_strand(session: Session, section: Section):
    track = session.get(Track, section.track_id) if section.track_id else None
    strand = session.get(Strand, section.strand_id) if section.strand_id else None
    return track, strand


def _fill_header(session, worksheet, anchors, *, school, section, school_year, term, track, strand,
                 year, month, class_day_count, adviser):
    write_ref(worksheet, anchors, CELL_SCHOOL_NAME, school.school_name if school else "")
    write_ref(worksheet, anchors, CELL_SCHOOL_ID, school.deped_school_id if school else "")
    write_ref(worksheet, anchors, CELL_DISTRICT, getattr(school, "district", "") or "")
    write_ref(worksheet, anchors, CELL_DIVISION, school.schools_division if school else "")
    write_ref(worksheet, anchors, CELL_REGION, school.region if school else "")
    write_ref(worksheet, anchors, CELL_TERM, term.name if term else "")
    write_ref(worksheet, anchors, CELL_SCHOOL_YEAR, school_year.name if school_year else "")
    write_ref(worksheet, anchors, CELL_GRADE_LEVEL, _grade_level_label(session, section))
    write_ref(
        worksheet,
        anchors,
        CELL_TRACK_STRAND,
        " / ".join(
            part for part in [track.name if track else "", strand.name if strand else ""] if part
        ),
    )
    write_ref(worksheet, anchors, CELL_MONTH, _calendar.month_name[month])
    write_ref(worksheet, anchors, CELL_SECTION, section.name)
    write_ref(worksheet, anchors, CELL_DAYS_OF_CLASSES, f"No. of Days of Classes: {class_day_count}")
    write_ref(worksheet, anchors, CELL_ADVISER, adviser.full_name.upper() if adviser else "")
    write_ref(worksheet, anchors, CELL_SCHOOL_HEAD, (school.school_head_name or "").upper() if school else "")


def _grade_level_label(session: Session, section) -> str:
    """Looked up explicitly — this codebase declares no ORM
    relationship(), so `section.grade_level` does not exist."""
    if section.grade_level_id is None:
        return ""
    level = session.get(GradeLevel, section.grade_level_id)
    return (level.code or level.name) if level else ""


def _fill_day_headers(worksheet, anchors, class_days) -> None:
    """Writes the day-of-month number and weekday letter for each class
    day, blanks any unused day slot so a short month doesn't inherit the
    template's leftovers, and clears the date scaffolding row.

    The scaffold row is always cleared, never written: it sits above the
    table header, so a populated one prints raw dates over the form.
    """
    for index, column in enumerate(DAY_COLS):
        write(worksheet, anchors, ROW_DATE_SCAFFOLD, column, None)
        if index < len(class_days):
            day = class_days[index].calendar_date
            write(worksheet, anchors, ROW_DAY_NUMBER, column, day.day)
            write(worksheet, anchors, ROW_WEEKDAY, column, WEEKDAY_LETTERS[day.weekday()])
        else:
            for row in (ROW_DAY_NUMBER, ROW_WEEKDAY):
                write(worksheet, anchors, row, column, None)


def _fill_learner_block(worksheet, anchors, rows, first_row: int, class_days) -> None:
    for offset in range(ROWS_PER_SEX):
        row_number = first_row + offset
        if offset < len(rows):
            entry = rows[offset]
            # Explicitly unhide: a sheet copied from an already-filled page
            # inherits its hidden flags.
            worksheet.row_dimensions[row_number].hidden = False
            write(worksheet, anchors, row_number, COL_NO, offset + 1)
            write(worksheet, anchors, row_number, COL_NAME, entry["name"])
            for index, column in enumerate(DAY_COLS):
                write(
                    worksheet,
                    anchors,
                    row_number,
                    column,
                    entry["marks"][index] if index < len(class_days) else None,
                )
            write(worksheet, anchors, row_number, COL_ABSENT, entry["absent"])
            write(worksheet, anchors, row_number, COL_PRESENT, entry["present"])
            write(worksheet, anchors, row_number, COL_REMARKS, entry["remarks"] or None)
        else:
            # Clear the leftover sample data and hide the row. Hiding is
            # what keeps unused rows off both the Excel print-out and the
            # PDF — `app/xlsx_render.py` gives a hidden row zero height,
            # so one mechanism covers both exports.
            worksheet.row_dimensions[row_number].hidden = True
            write(worksheet, anchors, row_number, COL_NO, offset + 1)
            write(worksheet, anchors, row_number, COL_NAME, None)
            for column in DAY_COLS:
                write(worksheet, anchors, row_number, column, None)
            for column in (COL_ABSENT, COL_PRESENT, COL_REMARKS):
                write(worksheet, anchors, row_number, column, None)


def _present_on(entry, index: int) -> bool:
    """A learner counts toward a day's total when the day is inside their
    window and they weren't marked absent. Blank means present here (the
    printed convention), and LATE/CUTTING still count as present."""
    mark = entry["marks"][index]
    return mark != "X"


def _fill_daily_totals(worksheet, anchors, males, females, class_days) -> None:
    for index, column in enumerate(DAY_COLS):
        if index >= len(class_days):
            for row in (ROW_MALE_TOTAL, ROW_FEMALE_TOTAL, ROW_COMBINED_TOTAL):
                write(worksheet, anchors, row, column, None)
            continue
        day = class_days[index].calendar_date
        male_total = sum(
            1 for e in males if e["window"].contains(day) and _present_on(e, index)
        )
        female_total = sum(
            1 for e in females if e["window"].contains(day) and _present_on(e, index)
        )
        write(worksheet, anchors, ROW_MALE_TOTAL, column, male_total)
        write(worksheet, anchors, ROW_FEMALE_TOTAL, column, female_total)
        write(worksheet, anchors, ROW_COMBINED_TOTAL, column, male_total + female_total)


def _write_summary_row(worksheet, anchors, row: int, male: float, female: float, total=None) -> None:
    """Total defaults to male + female, which is right for counts and for
    average daily attendance. It is NOT right for the percentage rows —
    adding 100% and 100% would report 200% — so those pass their own
    total, recomputed from the combined figures.
    """
    computed_total = male + female if total is None else total
    if isinstance(computed_total, float):
        # Averages summed as floats otherwise land as 2.7800000000000002.
        computed_total = round(computed_total, 4)
    write(worksheet, anchors, row, COL_SUMMARY_M, male)
    write(worksheet, anchors, row, COL_SUMMARY_F, female)
    write(worksheet, anchors, row, COL_SUMMARY_TOTAL, computed_total)


def _fill_summary(session, worksheet, anchors, males, females, class_days, school_year, year, month) -> None:
    day_count = len(class_days)

    registered_male = len([e for e in males if _registered_at_month_end(e, year, month)])
    registered_female = len([e for e in females if _registered_at_month_end(e, year, month)])

    start_reference = (
        first_friday_on_or_after(school_year.start_date) if school_year else None
    )
    enrolled_male = len([e for e in males if _active_on(e, start_reference)])
    enrolled_female = len([e for e in females if _active_on(e, start_reference)])

    male_attendance = sum(e["present"] for e in males)
    female_attendance = sum(e["present"] for e in females)
    average_male = round(male_attendance / day_count, 2) if day_count else 0
    average_female = round(female_attendance / day_count, 2) if day_count else 0

    _write_summary_row(worksheet, anchors, ROW_ENROLMENT_FIRST_FRIDAY, enrolled_male, enrolled_female)
    _write_summary_row(
        worksheet,
        anchors,
        ROW_LATE_ENROLMENT,
        len([e for e in males if _late_enrolled(e, year, month)]),
        len([e for e in females if _late_enrolled(e, year, month)]),
    )
    _write_summary_row(worksheet, anchors, ROW_REGISTERED_END, registered_male, registered_female)
    _write_summary_row(
        worksheet,
        anchors,
        ROW_PCT_ENROLMENT,
        _ratio(registered_male, enrolled_male),
        _ratio(registered_female, enrolled_female),
        total=_ratio(registered_male + registered_female, enrolled_male + enrolled_female),
    )
    _write_summary_row(worksheet, anchors, ROW_AVG_DAILY_ATTENDANCE, average_male, average_female)
    _write_summary_row(
        worksheet,
        anchors,
        ROW_PCT_ATTENDANCE,
        _ratio(average_male, registered_male),
        _ratio(average_female, registered_female),
        total=_ratio(average_male + average_female, registered_male + registered_female),
    )
    _write_summary_row(
        worksheet,
        anchors,
        ROW_FIVE_CONSECUTIVE,
        len([e for e in males if e["summary"].has_consecutive_absence_warning]),
        len([e for e in females if e["summary"].has_consecutive_absence_warning]),
    )

    counts = _movement_counts(session, males + females, year, month)
    for movement_type, row in (
        (EnrollmentStatus.NLS, ROW_NLS),
        (EnrollmentStatus.TRANSFERRED_OUT, ROW_TRANSFERRED_OUT),
        (EnrollmentStatus.TRANSFERRED_IN, ROW_TRANSFERRED_IN),
        (EnrollmentStatus.SHIFTED_OUT, ROW_SHIFTED_OUT),
        (EnrollmentStatus.SHIFTED_IN, ROW_SHIFTED_IN),
    ):
        bucket = counts.get(movement_type, {"M": 0, "F": 0})
        _write_summary_row(worksheet, anchors, row, bucket["M"], bucket["F"])


def _ratio(numerator, denominator):
    """The template formats these cells as percentages, so a fraction is
    what belongs in the cell — 1.0 renders as 100%."""
    return round(numerator / denominator, 4) if denominator else 0


def _active_on(entry, day: date | None) -> bool:
    return bool(day) and entry["window"].contains(day)


def _registered_at_month_end(entry, year: int, month: int) -> bool:
    """Still on the roll at the close of the month — a learner who
    transferred out mid-month appears on the sheet (§32) but is no longer
    registered for the end-of-month count."""
    month_end = date(year, month, _calendar.monthrange(year, month)[1])
    window = entry["window"]
    return window.end is None or window.end >= month_end


def _late_enrolled(entry, year: int, month: int) -> bool:
    window = entry["window"]
    if window.start is None:
        return False
    return window.start.year == year and window.start.month == month
