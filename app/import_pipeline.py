"""Migration import from the existing Excel workbook (§51).

**Never import silently.** §51 mandates the full sequence — upload,
detect columns, map columns, validate, show errors, confirm, and leave an
audit record — and this module is that sequence, parameterised by an
`ImportSpec` so each importable kind supplies only its own columns,
validators and writer.

The pipeline is deliberately split so the parts that don't need a
database can be tested without one:

  read_table()      bytes  -> headers + rows          (no DB)
  suggest_mapping() headers -> {field: column}        (no DB)
  validate()        rows   -> parsed rows + errors    (DB, batched)
  commit()          parsed -> rows written            (DB)

Reference lookups inside `validate` are loaded **once** for the whole
file rather than per row. The database is ~85ms away, so a per-row query
would make a 300-row masterlist take half a minute — the same trap that
made the report pages slow.
"""

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Callable

import openpyxl

from app.models.admin import ImportJob
from app.models.enums import ImportJobStatus, ImportJobType

# A row number the user will recognise: spreadsheet rows are 1-based and
# row 1 is the header, so the first data row is 2.
FIRST_DATA_ROW = 2


@dataclass
class ColumnSpec:
    """One importable field. `aliases` are the header spellings seen in
    the school's existing workbook, used to pre-select the mapping so the
    common case needs no manual work."""

    field: str
    label: str
    required: bool = True
    aliases: tuple[str, ...] = ()


@dataclass
class RowError:
    row_number: int
    column: str | None
    message: str

    def as_dict(self) -> dict:
        return {"row": self.row_number, "column": self.column, "message": self.message}


@dataclass
class ValidationResult:
    parsed: list[dict] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    def error_dicts(self) -> list[dict]:
        return [e.as_dict() for e in self.errors]


@dataclass
class ImportSpec:
    """Everything specific to one importable kind."""

    job_type: ImportJobType
    label: str
    description: str
    columns: list[ColumnSpec]
    # (session, rows, mapping) -> ValidationResult. Specs that set
    # `needs_school_year` also take a `school_year_id` keyword.
    validate: Callable
    # (session, parsed_rows, user_id) -> int rows written
    commit: Callable
    # Whether the page must ask which school year the file belongs to.
    # Kept off the file itself: a school year repeated on all 1,200 rows
    # is 1,200 chances to disagree with itself.
    needs_school_year: bool = False

    @property
    def required_fields(self) -> list[str]:
        return [c.field for c in self.columns if c.required]


# --- Reading (no database) -------------------------------------------------


def _normalise(header: str) -> str:
    """Collapses a header to a comparable key so "Last Name", "LAST_NAME"
    and "lastname" all match."""
    return re.sub(r"[^a-z0-9]", "", str(header or "").lower())


def read_table(data: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Returns (headers, rows) from an .xlsx or .csv upload.

    Every value comes back as a **string**, untouched. Parsing into dates
    and numbers is the validator's job, so that a bad value produces a
    row-level error the user can act on rather than an exception here.
    Reading LRNs as text also stops Excel's numeric coercion from eating
    a leading zero before validation ever sees it.
    """
    if filename.lower().endswith((".xlsx", ".xlsm")):
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        raw = [list(row) for row in worksheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        text = data.decode("utf-8-sig", errors="replace")
        raw = [row for row in csv.reader(io.StringIO(text))]

    raw = [row for row in raw if any(_cell(v) for v in row)]
    if not raw:
        return [], []

    headers = [_cell(v) for v in raw[0]]
    rows = []
    for offset, values in enumerate(raw[1:], start=FIRST_DATA_ROW):
        row = {headers[i]: _cell(v) for i, v in enumerate(values) if i < len(headers)}
        row["__row__"] = offset
        rows.append(row)
    return headers, rows


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # Excel hands back 12-digit LRNs as floats; str() would render
        # 1.07041140016e+11 and lose the value entirely.
        return str(int(value))
    return str(value).strip()


def suggest_mapping(headers: list[str], spec: ImportSpec) -> dict[str, str]:
    """Best-guess {field: header}, matching on a normalised header and the
    spec's known aliases. Anything unmatched is left for the user."""
    by_key = {_normalise(h): h for h in headers if h}
    mapping: dict[str, str] = {}
    for column in spec.columns:
        candidates = (column.field, column.label, *column.aliases)
        for candidate in candidates:
            header = by_key.get(_normalise(candidate))
            if header:
                mapping[column.field] = header
                break
    return mapping


def missing_required(mapping: dict[str, str], spec: ImportSpec) -> list[str]:
    return [c.label for c in spec.columns if c.required and not mapping.get(c.field)]


def apply_mapping(rows: list[dict], mapping: dict[str, str]) -> list[dict]:
    """Rewrites raw rows keyed by spreadsheet header into rows keyed by
    spec field, keeping the original row number for error reporting."""
    mapped = []
    for row in rows:
        item = {field: row.get(header, "") for field, header in mapping.items() if header}
        item["__row__"] = row.get("__row__")
        mapped.append(item)
    return mapped


# --- Shared value parsing --------------------------------------------------


def parse_grade(raw: str) -> tuple[int | None, str | None]:
    """A grade, or an explanation. Blank is allowed and means *not yet
    encoded* (rule 2) — it must never become a zero."""
    if raw in ("", None):
        return None, None
    try:
        value = float(str(raw).replace(",", "."))
    except ValueError:
        return None, f"{raw!r} is not a number"
    if not (0 <= value <= 100):
        return None, f"{raw} is outside 0-100"
    return int(round(value)), None


# Which way round a slash date is read when the file gives no evidence.
# Excel writes whatever the PC's regional setting says, and the machines
# this runs on are set to US order — a birthdate typed as 2009-01-15 comes
# back out of Excel as 01/15/2009.
DEFAULT_DATE_ORDER = "mdy"

_SLASH_DATE = re.compile(r"\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})\s*")


def detect_date_order(values) -> str:
    """Decide whether a column of slash dates is month-first or day-first.

    03/04/2009 is genuinely two different days and no amount of care can
    tell them apart in isolation. A *column* usually can: one value with a
    first component above 12 proves day-first, one with a second component
    above 12 proves month-first. Only when the whole column is ambiguous
    does this fall back to the regional default.

    Getting this wrong is silent — a learner is simply born on the wrong
    day, and nothing ever errors — so it is decided once for the file
    rather than guessed per row.
    """
    first_over_12 = second_over_12 = False
    for raw in values:
        match = _SLASH_DATE.fullmatch(str(raw or ""))
        if not match:
            continue
        first, second = int(match.group(1)), int(match.group(2))
        first_over_12 |= first > 12
        second_over_12 |= second > 12
    if first_over_12 and not second_over_12:
        return "dmy"
    if second_over_12 and not first_over_12:
        return "mdy"
    return DEFAULT_DATE_ORDER


def parse_date(raw: str, order: str = DEFAULT_DATE_ORDER) -> tuple[date | None, str | None]:
    """§51 calls out "impossible date" as a validation error, so an
    unparseable or out-of-range value is reported rather than raised.

    `order` comes from `detect_date_order` over the whole column. ISO is
    always tried first and is never ambiguous, so a file that was written
    as YYYY-MM-DD is unaffected by any of this.
    """
    if raw in ("", None):
        return None, None
    slash = ("%d/%m/%Y", "%m/%d/%Y") if order == "dmy" else ("%m/%d/%Y", "%d/%m/%Y")
    patterns = ("%Y-%m-%d", "%Y/%m/%d", *slash, "%d-%b-%Y", "%d %B %Y", "%B %d, %Y")
    for pattern in patterns:
        try:
            return datetime.strptime(str(raw).strip(), pattern).date(), None
        except ValueError:
            continue
    return None, f"{raw!r} is not a date we can read (try YYYY-MM-DD)"


def parse_lrn(raw: str) -> tuple[str | None, str | None]:
    """LRN stays text so leading zeros survive (rule 10).

    Excel treats a 12-digit LRN as a number, and the two ways that damages
    it are both recoverable here rather than by asking a teacher to
    reformat the column as text before uploading:

    - a cell read as a float arrives as "107041140016.0"
    - a cell Excel saved to CSV in its display form arrives as "1.07041E+11"

    Both are undone arithmetically, so nothing is guessed. A value that is
    genuinely the wrong length is still rejected — the point is to stop
    Excel corrupting a correct number, not to accept an incorrect one.
    """
    value = str(raw or "").strip().replace(" ", "")
    if not value:
        return None, None

    if re.fullmatch(r"\d+\.0+", value):
        value = value.split(".")[0]
    elif re.fullmatch(r"\d(\.\d+)?[Ee]\+?\d+", value):
        try:
            expanded = Decimal(value)
        except InvalidOperation:
            return None, f"LRN {str(raw).strip()!r} is not a number we can read"
        digits = str(int(expanded))
        # Only trust the expansion if the mantissa still carries every
        # digit. "1.07041140016E+11" does; "1.07E+11" does not — Excel
        # rounded it away when it wrote the file, and expanding that
        # would invent 107000000000, a number the learner does not have.
        # A fabricated LRN is far worse than a rejected row.
        significant = len(re.sub(r"[^0-9]", "", value.split("E")[0].split("e")[0]).lstrip("0"))
        if expanded != expanded.to_integral_value() or significant < len(digits):
            return None, (
                f"LRN {str(raw).strip()!r} has been rounded by Excel and its digits are "
                "lost — widen the column or format it as text, then save and re-upload"
            )
        value = digits

    if not value.isdigit() or len(value) != 12:
        return None, f"LRN {str(raw).strip()!r} must be exactly 12 digits"
    return value, None


# --- Job records (the audit trail §51 step 7) ------------------------------


def create_job(session, spec: ImportSpec, filename: str, user_id, row_count: int) -> ImportJob:
    job = ImportJob(
        job_type=spec.job_type,
        status=ImportJobStatus.UPLOADED,
        uploaded_by_user_id=user_id,
        file_path=filename,  # the uploaded name; the bytes aren't retained
        row_count=row_count,
    )
    session.add(job)
    session.flush()
    return job


def record_validation(session, job: ImportJob, mapping: dict, result: ValidationResult) -> None:
    job.column_mapping = mapping
    job.validation_errors = result.error_dicts() or None
    job.status = ImportJobStatus.VALIDATED if result.ok else ImportJobStatus.FAILED
    session.flush()


def record_confirmation(session, job: ImportJob, imported: int) -> None:
    job.status = ImportJobStatus.CONFIRMED
    job.imported_count = imported
    job.confirmed_at = datetime.now(timezone.utc)
    session.flush()
