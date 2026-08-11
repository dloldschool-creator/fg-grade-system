"""Render a filled .xlsx worksheet straight to PDF with ReportLab.

**Why this exists.** The original pipeline (CLAUDE.md's stack decision)
filled the official DepEd template with openpyxl and flattened it with
headless LibreOffice. That works, but LibreOffice costs ~261 MB of RSS per
conversion with no concurrency limit, which is the single largest memory
risk in the deployed app. This module removes that dependency.

**The rule it follows, and the reason it's a renderer rather than a
hand-drawn form.** §56 requires DATA and PRINT TEMPLATE to stay separate,
so a DepEd revision means swapping a file, not rewriting code. Nothing
here knows what an SF9 *is*: it reads the geometry the template itself
carries -- merges, column widths, row heights, borders, fonts, images,
page setup -- and draws that. Drop in a revised template, or SF10 when it
arrives, and this renders it with no changes.

**Coordinate systems.** Excel addresses cells row-down from the top left
and sizes columns in "characters" of the default font; ReportLab draws
from the bottom left in points. Everything is converted to a top-left
points space first (`sheet_geometry`), then flipped once at draw time, so
no other function has to think about it.
"""

import io
import os
from dataclasses import dataclass

from openpyxl.utils import column_index_from_string, get_column_letter
from reportlab.lib.pagesizes import letter, legal
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdfcanvas

# --- Unit conversion -------------------------------------------------------

EMU_PER_POINT = 12700
PIXELS_TO_POINTS = 0.75  # 96 dpi screen pixels -> 72 dpi points

# Excel stores a column width as a count of "0" glyphs in the workbook's
# default font. For the 11pt fonts these templates use, one such glyph is
# 7px wide and the cell carries 5px of padding -- the conversion Excel's
# own file-format documentation gives.
MAX_DIGIT_WIDTH = 7
CELL_PADDING_PX = 5

DEFAULT_COL_WIDTH = 8.43   # Excel's own default, in those glyph units
DEFAULT_ROW_HEIGHT = 15.0  # points

# Drawn inset so text doesn't touch its cell border, matching Excel's own
# small horizontal padding.
TEXT_INSET = 2.0

# Slack when deciding whether one more row still fits on the page. See
# plan_pages — without it, a form scaled to exactly fill its page spills a
# single row onto a second one.
PAGE_BREAK_TOLERANCE = 0.5


def column_width_to_points(width: float | None, *, is_stored: bool = True) -> float:
    """Excel column width to points.

    **The padding is already inside a stored width, and adding it again is
    a real bug.** ECMA-376 defines the stored value as

        width = Truncate((chars * MDW + 5) / MDW * 256) / 256

    so it is chars + 5/MDW, and inverting gives `pixels = width * MDW`
    with no further padding. Adding 5px per column on top inflated SF2 —
    81 columns, many of them sub-character hairline separators — from
    10.3in to 14.6in, which shrank the whole form to 62% and printed its
    title at 6.9pt against the 10pt it should be.

    `baseColWidth` is the exception: it *is* a raw character count, so the
    padding does apply there.
    """
    if width is None:
        width = DEFAULT_COL_WIDTH
        is_stored = False
    pixels = width * MAX_DIGIT_WIDTH + (0 if is_stored else CELL_PADDING_PX)
    return pixels * PIXELS_TO_POINTS


# --- Fonts -----------------------------------------------------------------

# The templates name Arial, Arial Narrow, Carlito, SansSerif and Aparajita.
# ReportLab ships only the base-14 families, so anything else needs a TTF.
# Arial in particular is not freely redistributable, so the default is a
# metric-compatible substitute rather than shipping the real file.
#
# Worth knowing: the old LibreOffice pipeline had this same substitution
# happening invisibly -- it only looked exact locally because Windows has
# these fonts installed. A Linux server substitutes either way.
_FONT_SUBSTITUTIONS = {
    "arial": "Helvetica",
    "arial narrow": "Helvetica",
    "helvetica": "Helvetica",
    "calibri": "Helvetica",
    "carlito": "Helvetica",
    "sansserif": "Helvetica",
    "aparajita": "Helvetica",
    "times new roman": "Times-Roman",
    "cambria": "Times-Roman",
    "courier new": "Courier",
}

_BASE14_VARIANTS = {
    "Helvetica": ("Helvetica", "Helvetica-Bold", "Helvetica-Oblique", "Helvetica-BoldOblique"),
    "Times-Roman": ("Times-Roman", "Times-Bold", "Times-Italic", "Times-BoldItalic"),
    "Courier": ("Courier", "Courier-Bold", "Courier-Oblique", "Courier-BoldOblique"),
}

_registered: dict[str, str] = {}


def register_font_file(family: str, path: str, *, bold: bool = False, italic: bool = False) -> None:
    """Registers a real TTF so a form can print in its actual typeface.

    Optional by design: without it everything falls back to a
    metric-compatible base-14 face, which keeps the layout correct even on
    a server with no fonts installed."""
    if not os.path.exists(path):
        return
    suffix = ("-Bold" if bold else "") + ("-Italic" if italic else "")
    name = f"{family}{suffix}"
    pdfmetrics.registerFont(TTFont(name, path))
    _registered[(family.lower(), bold, italic)] = name


def resolve_font(name: str | None, bold: bool, italic: bool) -> str:
    key = (name or "").strip().lower()
    explicit = _registered.get((key, bold, italic))
    if explicit:
        return explicit
    base = _FONT_SUBSTITUTIONS.get(key, "Helvetica")
    regular, b, i, bi = _BASE14_VARIANTS[base]
    if bold and italic:
        return bi
    if bold:
        return b
    if italic:
        return i
    return regular


# --- Geometry --------------------------------------------------------------


@dataclass
class Geometry:
    """Cell edges in a top-left-origin points space."""

    col_x: list[float]   # left edge of each column, 1-indexed with a [0] pad
    col_w: list[float]
    row_y: list[float]   # top edge of each row
    row_h: list[float]
    max_row: int
    max_col: int

    @property
    def width(self) -> float:
        return self.col_x[self.max_col] + self.col_w[self.max_col]

    @property
    def height(self) -> float:
        return self.row_y[self.max_row] + self.row_h[self.max_row]


def column_dimension_map(worksheet, max_col: int) -> dict[int, object]:
    """Column index -> its dimension, expanding grouped definitions.

    A worksheet stores widths as `<col min= max= width=>` **ranges**, and
    openpyxl keys them by the first letter only: SF9 has one entry under
    "P" that actually governs columns 16-26. Looking them up per letter
    silently misses every column but the first of each range, which then
    falls back to the default width — on SF9 that inflated the form from
    12.25in wide to 16.11in and shrank the whole card to 65% to make it
    fit.
    """
    by_index: dict[int, object] = {}
    for key, dim in worksheet.column_dimensions.items():
        start = getattr(dim, "min", None)
        end = getattr(dim, "max", None)
        if start is None:
            # A dimension created in memory has no min/max until the
            # workbook is saved; its dict key is still the column letter.
            try:
                start = end = column_index_from_string(key)
            except (ValueError, AttributeError):
                continue
        for col in range(start, min(end or start, max_col) + 1):
            by_index[col] = dim
    return by_index


def sheet_geometry(worksheet, max_row: int | None = None, max_col: int | None = None) -> Geometry:
    max_row = max_row or worksheet.max_row
    max_col = max_col or worksheet.max_column

    # baseColWidth is the sheet's own default when defaultColWidth is unset;
    # only fall back to Excel's global 8.43 when neither is present.
    fmt = worksheet.sheet_format
    # defaultColWidth is a stored (already-padded) value; baseColWidth is a
    # raw character count. They need different conversions — see
    # column_width_to_points.
    default_col = getattr(fmt, "defaultColWidth", None)
    default_is_stored = default_col is not None
    if default_col is None:
        default_col = getattr(fmt, "baseColWidth", None) or DEFAULT_COL_WIDTH
    default_row = getattr(fmt, "defaultRowHeight", None) or DEFAULT_ROW_HEIGHT

    dimensions = column_dimension_map(worksheet, max_col)
    col_w = [0.0] * (max_col + 1)
    col_x = [0.0] * (max_col + 1)
    x = 0.0
    for col in range(1, max_col + 1):
        dim = dimensions.get(col)
        # A hidden column occupies no space on the printed page.
        if dim is not None and dim.hidden:
            width = 0.0
        elif dim is not None and dim.width:
            width = column_width_to_points(dim.width, is_stored=True)
        else:
            width = column_width_to_points(default_col, is_stored=default_is_stored)
        col_x[col] = x
        col_w[col] = width
        x += width

    row_h = [0.0] * (max_row + 1)
    row_y = [0.0] * (max_row + 1)
    y = 0.0
    for row in range(1, max_row + 1):
        dim = worksheet.row_dimensions.get(row)
        if dim is not None and dim.hidden:
            height = 0.0
        else:
            height = float(dim.height) if dim is not None and dim.height else float(default_row)
        row_y[row] = y
        row_h[row] = height
        y += height

    return Geometry(col_x, col_w, row_y, row_h, max_row, max_col)


def merged_spans(worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """(row, col) of each merge anchor -> (rowspan, colspan)."""
    spans = {}
    for rng in worksheet.merged_cells.ranges:
        spans[(rng.min_row, rng.min_col)] = (
            rng.max_row - rng.min_row + 1,
            rng.max_col - rng.min_col + 1,
        )
    return spans


def covered_cells(worksheet) -> set[tuple[int, int]]:
    """Cells swallowed by a merge — drawn as part of their anchor, never
    on their own."""
    covered = set()
    for rng in worksheet.merged_cells.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if (row, col) != (rng.min_row, rng.min_col):
                    covered.add((row, col))
    return covered


# --- Drawing ---------------------------------------------------------------

_BORDER_WIDTHS = {
    "hair": 0.25,
    "thin": 0.5,
    "medium": 1.0,
    "thick": 1.5,
    "double": 0.5,
    "dotted": 0.5,
    "dashed": 0.5,
    "dashDot": 0.5,
    "dashDotDot": 0.5,
    "mediumDashed": 1.0,
    "slantDashDot": 0.5,
}

_DASH_PATTERNS = {
    "dotted": (1, 2),
    "dashed": (3, 2),
    "dashDot": (3, 2, 1, 2),
    "dashDotDot": (3, 2, 1, 2, 1, 2),
    "mediumDashed": (4, 2),
}


def _rgb(color) -> tuple[float, float, float] | None:
    """openpyxl colours arrive as ARGB hex, or as theme/indexed references
    this renderer deliberately doesn't resolve — those return None so the
    caller can fall back to a sensible default rather than guess wrong."""
    if color is None or color.type != "rgb" or not color.rgb:
        return None
    value = str(color.rgb)
    if len(value) == 8:
        value = value[2:]  # drop alpha
    if len(value) != 6:
        return None
    return tuple(int(value[i:i + 2], 16) / 255 for i in (0, 2, 4))


def _wrap(text: str, font: str, size: float, max_width: float) -> list[str]:
    """Word-wraps while preserving leading whitespace.

    That indent is load-bearing, not cosmetic: `report_card` marks a
    combined-language component row with ten leading spaces
    (`COMPONENT_INDENT`), which is what shows a parent learning area and
    its two components as a hierarchy on the SF9 (§16). A plain
    `text.split()` throws it away and the card reads as three unrelated
    subjects.
    """
    if max_width <= 0:
        return [text]
    indent = text[: len(text) - len(text.lstrip())]
    indent_width = pdfmetrics.stringWidth(indent, font, size) if indent else 0.0

    lines, current = [], ""
    for word in text.split():
        candidate = f"{current} {word}".strip()
        allowed = max_width - (indent_width if not lines else 0)
        if pdfmetrics.stringWidth(candidate, font, size) <= allowed or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    if not lines:
        return [text]
    # Only the first line carries the indent — a wrapped continuation
    # lining up under it would read as a further level of nesting.
    lines[0] = indent + lines[0]
    return lines


def _draw_cell_fill(c, cell, x, y, w, h) -> None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return
    rgb = _rgb(fill.fgColor)
    if rgb is None:
        return
    c.setFillColorRGB(*rgb)
    c.rect(x, y - h, w, h, stroke=0, fill=1)


def _draw_cell_borders(c, cell, x, y, w, h) -> None:
    border = cell.border
    if border is None:
        return
    edges = (
        ("top", x, y, x + w, y),
        ("bottom", x, y - h, x + w, y - h),
        ("left", x, y, x, y - h),
        ("right", x + w, y, x + w, y - h),
    )
    for name, x1, y1, x2, y2 in edges:
        side = getattr(border, name, None)
        if side is None or not side.style:
            continue
        c.setLineWidth(_BORDER_WIDTHS.get(side.style, 0.5))
        rgb = _rgb(side.color) or (0, 0, 0)
        c.setStrokeColorRGB(*rgb)
        dash = _DASH_PATTERNS.get(side.style)
        c.setDash(*dash) if dash else c.setDash()
        c.line(x1, y1, x2, y2)
        if side.style == "double":
            offset = 1.2
            dx = 0 if name in ("top", "bottom") else offset
            dy = offset if name in ("top", "bottom") else 0
            c.line(x1 + dx, y1 - dy, x2 + dx, y2 - dy)
    c.setDash()


def apply_number_format(value, number_format: str | None) -> str:
    """Applies the subset of Excel number formats these forms actually use.

    Not decorative. SF2's summary box stores a percentage as the ratio 1.0
    with a `0.00%` format; printing the raw value puts "1" where the
    division office expects "100.00%". Excel's format language is far
    larger than this, so anything unrecognised falls through to a plain
    string rather than being guessed at.
    """
    if value is None:
        return ""
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return str(value)

    fmt = (number_format or "General").split(";")[0].strip()
    if fmt in ("General", "@", ""):
        return str(int(value)) if float(value).is_integer() else str(value)

    percent = fmt.endswith("%")
    body = fmt[:-1] if percent else fmt
    if percent:
        value = value * 100

    thousands = "," in body

    # `0` is a required digit, `#` and `?` are optional ones that suppress
    # trailing zeros — so `#.###` shows 1 as "1", not "1.000", while
    # `0.00` shows it as "1.00". Treating them alike put "1.000" in SF2's
    # No. column. Only the placeholder run immediately after the point
    # counts; an accounting format like `_(* #,##0.00_)` trails literals.
    minimum = maximum = 0
    if "." in body:
        for character in body.split(".", 1)[1]:
            if character == "0":
                minimum += 1
                maximum += 1
            elif character in "#?":
                maximum += 1
            else:
                break

    try:
        text = f"{value:,.{maximum}f}" if thousands else f"{value:.{maximum}f}"
    except (ValueError, TypeError):
        return str(value)

    if maximum > minimum and "." in text:
        whole, _, fraction = text.partition(".")
        fraction = fraction.rstrip("0")
        while len(fraction) < minimum:
            fraction += "0"
        text = f"{whole}.{fraction}" if fraction else whole

    return text + "%" if percent else text


def _format_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        # A formula that was never evaluated prints as its own source text
        # otherwise, which is worse than a blank on an official form.
        return "" if value.startswith("=") else value
    return apply_number_format(value, cell.number_format)


def overflow_width(worksheet, geometry, row, col, colspan, covered) -> float:
    """How far a left-aligned label may run before it hits something.

    Excel lets text spill across empty neighbouring cells rather than
    shrinking it — which is why "Dear Parent/Guardian," and "Eligible for
    Admission to Grade:" sit in one narrow column on these forms and still
    read at full size. Without this the renderer shrinks them to fit their
    own column and the form looks subtly wrong everywhere.
    """
    width = sum(geometry.col_w[col:col + colspan])
    next_col = col + colspan
    while next_col <= geometry.max_col:
        if (row, next_col) in covered:
            break
        neighbour = worksheet.cell(row=row, column=next_col)
        if neighbour.value not in (None, ""):
            break
        width += geometry.col_w[next_col]
        next_col += 1
    return width


def _draw_cell_text(c, cell, x, y, w, h, *, overflow: float | None = None) -> None:
    text = _format_value(cell)
    if not text:
        return

    font_obj = cell.font
    size = float(font_obj.sz or 11)
    font = resolve_font(font_obj.name, bool(font_obj.b), bool(font_obj.i))
    rgb = _rgb(font_obj.color) or (0, 0, 0)

    alignment = cell.alignment
    horizontal = alignment.horizontal or ("right" if isinstance(cell.value, (int, float)) else "left")
    vertical = alignment.vertical or "bottom"

    usable = w - 2 * TEXT_INSET
    lines = _wrap(text, font, size, usable) if alignment.wrap_text else [text]

    # Only shrink once the text has exhausted the empty cells it may spill
    # into — and only for left/general alignment, since a centred or
    # right-aligned entry doesn't overflow the same way.
    if not alignment.wrap_text and usable > 0:
        spill = (overflow - 2 * TEXT_INSET) if (overflow and horizontal in ("left", "general", None)) else usable
        actual = pdfmetrics.stringWidth(text, font, size)
        if actual > max(usable, spill):
            size = max(size * max(usable, spill) / actual, 3.5)

    leading = size * 1.18
    block = leading * len(lines)
    if vertical == "top":
        cursor = y - size
    elif vertical == "center":
        cursor = y - (h - block) / 2 - size
    else:
        cursor = y - h + block - size + (leading - size) / 2

    c.setFillColorRGB(*rgb)
    c.setFont(font, size)
    for line in lines:
        line_width = pdfmetrics.stringWidth(line, font, size)
        if horizontal == "center" or horizontal == "centerContinuous":
            tx = x + (w - line_width) / 2
        elif horizontal == "right":
            tx = x + w - TEXT_INSET - line_width
        else:
            tx = x + TEXT_INSET
        c.drawString(tx, cursor, line)
        cursor -= leading


def _draw_images(c, worksheet, geometry, flip, first_row: int = 1, last_row: int | None = None) -> None:
    """Places the DepEd shield and the school seal.

    An anchor gives a cell plus an EMU offset, so the position has to be
    resolved through the same geometry the cells use — otherwise the seals
    drift as soon as a column width changes.
    """
    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        col = min(marker.col + 1, geometry.max_col)
        row = min(marker.row + 1, geometry.max_row)
        # Each seal is drawn once, on the page its anchor row falls in.
        if row < first_row or (last_row is not None and row > last_row):
            continue
        x = geometry.col_x[col] + marker.colOff / EMU_PER_POINT
        top = geometry.row_y[row] + marker.rowOff / EMU_PER_POINT

        # Three anchor shapes, and getting this wrong is spectacular rather
        # than subtle: a OneCellAnchor falling through to the image's own
        # natural pixel size renders a 1000px seal at ~10in wide, straight
        # across the header.
        to_marker = getattr(anchor, "to", None)
        extent = getattr(anchor, "ext", None)
        if to_marker is not None:
            end_col = min(to_marker.col + 1, geometry.max_col)
            end_row = min(to_marker.row + 1, geometry.max_row)
            width = (geometry.col_x[end_col] + to_marker.colOff / EMU_PER_POINT) - x
            height = (geometry.row_y[end_row] + to_marker.rowOff / EMU_PER_POINT) - top
        elif extent is not None:
            width = extent.cx / EMU_PER_POINT
            height = extent.cy / EMU_PER_POINT
        else:
            width = getattr(image, "width", 0) * PIXELS_TO_POINTS
            height = getattr(image, "height", 0) * PIXELS_TO_POINTS

        if width <= 0 or height <= 0:
            continue

        # The image's bytes can only be read once (the same openpyxl trap
        # excel_template.replicate_images works around), so take a copy.
        try:
            raw = image.ref
            data = raw.getvalue() if hasattr(raw, "getvalue") else open(raw, "rb").read()
        except Exception:
            continue
        try:
            c.drawImage(
                ImageReader(io.BytesIO(data)), x, flip(top) - height,
                width=width, height=height, mask="auto",
            )
        except Exception:
            continue  # a decorative seal must never break a report


# --- Page setup ------------------------------------------------------------

_PAGE_SIZES = {
    "letter": letter,
    "legal": legal,
    # DepEd forms are commonly printed on Philippine long bond, which is
    # 8.5 x 13in -- neither US Legal (14in) nor A4.
    "folio": (8.5 * 72, 13 * 72),
}


def page_size_for(worksheet, default=letter, landscape: bool | None = None):
    setup = worksheet.page_setup
    size = default
    if landscape is None:
        landscape = (setup.orientation == "landscape")
    if landscape:
        size = (max(size), min(size))
    else:
        size = (min(size), max(size))
    return size


def render_worksheet_to_pdf(
    worksheet,
    *,
    page=letter,
    landscape: bool | None = None,
    max_row: int | None = None,
    max_col: int | None = None,
    fit: bool = True,
) -> bytes:
    """Draws one worksheet onto a single page, scaled to fit.

    Fit-to-page is the default because the templates set fitToWidth and
    fitToHeight; a form that silently spills onto a second sheet is worse
    than one printed slightly small.
    """
    buffer = io.BytesIO()
    page_width, page_height = page_size_for(worksheet, page, landscape)
    c = pdfcanvas.Canvas(buffer, pagesize=(page_width, page_height))
    draw_worksheet(c, worksheet, page_width, page_height, max_row=max_row, max_col=max_col, fit=fit)
    c.showPage()
    c.save()
    return buffer.getvalue()


def page_count(workbook, *, page=letter, landscape: bool | None = None) -> int:
    """How many pages the workbook will print to, without rendering it."""
    total = 0
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue
        width, height = page_size_for(worksheet, page, landscape)
        _, bands = plan_pages(worksheet, width, height, sheet_geometry(worksheet))
        total += len(bands)
    return total


def _margins(worksheet) -> tuple[float, float, float, float]:
    m = worksheet.page_margins
    return (
        (m.left or 0.25) * 72, (m.right or 0.25) * 72,
        (m.top or 0.25) * 72, (m.bottom or 0.25) * 72,
    )


def plan_pages(worksheet, page_width: float, page_height: float, geometry: Geometry):
    """Decides the scale and the row bands each page carries.

    `fitToWidth` / `fitToHeight` are counts of pages, not booleans, and a
    zero means "as many as it takes". SF9 sets both to 1 and must land on
    a single sheet. **SF2 sets fitToWidth=1 and fitToHeight=0** — one page
    wide, unlimited pages tall — because a full 50-learner roster is
    15.4 x 22.6in. Squashing that onto one page needs 35% scale, which
    prints the form's 5pt text at 1.8pt: present, and unreadable.
    """
    left, right, top, bottom = _margins(worksheet)
    available_w = page_width - left - right
    available_h = page_height - top - bottom

    setup = worksheet.page_setup
    fit_enabled = bool(getattr(worksheet.sheet_properties.pageSetUpPr, "fitToPage", False))
    fit_w = setup.fitToWidth if setup.fitToWidth is not None else (1 if fit_enabled else 0)
    fit_h = setup.fitToHeight if setup.fitToHeight is not None else (1 if fit_enabled else 0)

    if not fit_enabled and setup.scale:
        scale = float(setup.scale) / 100
    else:
        candidates = []
        if fit_w:
            candidates.append(available_w * fit_w / geometry.width)
        if fit_h:
            candidates.append(available_h * fit_h / geometry.height)
        scale = min(candidates + [1.0]) if candidates else 1.0

    # Break on row boundaries, never through the middle of one.
    #
    # The tolerance matters: when fitToHeight pins the scale so the form
    # exactly fills the page, the scaled heights sum to a hair over the
    # limit in floating point and the final row spills to a second, nearly
    # empty page. SF9 fits in 575.94pt of 576 and did precisely that.
    # Half a point is far below anything visible and well under one row.
    bands: list[tuple[int, int]] = []
    start = 1
    used = 0.0
    for row in range(1, geometry.max_row + 1):
        height = geometry.row_h[row] * scale
        if used > 0 and used + height > available_h + PAGE_BREAK_TOLERANCE:
            bands.append((start, row - 1))
            start = row
            used = 0.0
        used += height
    bands.append((start, geometry.max_row))
    return scale, bands


def draw_worksheet(
    c,
    worksheet,
    page_width: float,
    page_height: float,
    *,
    max_row: int | None = None,
    max_col: int | None = None,
    fit: bool = True,
) -> int:
    """Draws the worksheet, starting a new page per row band. Returns the
    number of pages drawn; the caller owns the final `showPage`."""
    left, right, top, bottom = _margins(worksheet)
    geometry = sheet_geometry(worksheet, max_row, max_col)
    available_w = page_width - left - right
    available_h = page_height - top - bottom

    if fit:
        scale, bands = plan_pages(worksheet, page_width, page_height, geometry)
    else:
        scale, bands = 1.0, [(1, geometry.max_row)]

    spans = merged_spans(worksheet)
    covered = covered_cells(worksheet)
    options = worksheet.print_options
    content_w = geometry.width * scale

    for index, (first_row, last_row) in enumerate(bands):
        if index:
            c.showPage()

        band_top = geometry.row_y[first_row]
        band_height = (geometry.row_y[last_row] + geometry.row_h[last_row] - band_top) * scale

        offset_x = left + ((available_w - content_w) / 2 if options.horizontalCentered else 0)
        # Vertical centring applies to a single-page form; on a paginated
        # one it would float each band's rows away from the page edge and
        # make the break look like a layout error.
        centre_v = options.verticalCentered and len(bands) == 1
        offset_y = top + ((available_h - band_height) / 2 if centre_v else 0)

        c.saveState()
        c.translate(offset_x, page_height - offset_y)
        c.scale(scale, scale)

        def flip(top_y: float, _band_top=band_top) -> float:
            return -(top_y - _band_top)

        # Fills first, then borders, then text — so a border is never
        # painted over by the neighbouring cell's background.
        for pass_name in ("fill", "border", "text"):
            for row in range(first_row, last_row + 1):
                for col in range(1, geometry.max_col + 1):
                    if (row, col) in covered:
                        continue
                    rowspan, colspan = spans.get((row, col), (1, 1))
                    x = geometry.col_x[col]
                    y = flip(geometry.row_y[row])
                    w = sum(geometry.col_w[col:col + colspan])
                    h = sum(geometry.row_h[row:row + rowspan])
                    if w <= 0 or h <= 0:
                        continue
                    cell = worksheet.cell(row=row, column=col)
                    if pass_name == "fill":
                        _draw_cell_fill(c, cell, x, y, w, h)
                    elif pass_name == "border":
                        _draw_cell_borders(c, cell, x, y, w, h)
                    else:
                        _draw_cell_text(
                            c, cell, x, y, w, h,
                            overflow=overflow_width(worksheet, geometry, row, col, colspan, covered),
                        )

        # Images belong to the band whose rows they start in.
        _draw_images(c, worksheet, geometry, flip, first_row, last_row)
        c.restoreState()

    return len(bands)


def _visible_sheets(workbook):
    sheets = [ws for ws in workbook.worksheets if ws.sheet_state == "visible"]
    return sheets or list(workbook.worksheets)


def workbook_to_pdf(workbook, *, page=letter, landscape: bool | None = None, fit: bool = True) -> bytes:
    """Every visible worksheet — the replacement for the old
    LibreOffice conversion, with no external program involved."""
    return workbooks_to_pdf([workbook], page=page, landscape=landscape, fit=fit)


def workbooks_to_pdf(workbooks, *, page=letter, landscape: bool | None = None, fit: bool = True) -> bytes:
    """Several workbooks into one PDF — a whole section's report cards in
    a single file.

    Takes an iterable rather than a list so the caller can build each
    learner's workbook as it is drawn: 40 SF9s rendered this way cost one
    workbook of memory at a time instead of forty.
    """
    buffer = io.BytesIO()
    c = None
    for workbook in workbooks:
        for worksheet in _visible_sheets(workbook):
            width, height = page_size_for(worksheet, page, landscape)
            if c is None:
                c = pdfcanvas.Canvas(buffer, pagesize=(width, height))
            c.setPageSize((width, height))
            draw_worksheet(c, worksheet, width, height, fit=fit)
            c.showPage()
    if c is None:
        raise ValueError("nothing to render")
    c.save()
    return buffer.getvalue()
