"""Shared plumbing for filling the official DepEd Excel templates
(SF2, SF9, and SF10 later) with `openpyxl`.

Every one of those templates comes out of the same master automation
workbook on OneDrive, so they share the same three hazards — external
links, row-varying merge anchors, and images that don't survive a sheet
copy. Solving them once here keeps the per-form modules about the form's
own layout rather than about openpyxl.
"""

import io
from copy import copy

from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.formula import ArrayFormula

# Prefix Excel gives a reference into the first externally-linked
# workbook. Every data cell in these templates reads from the school's
# master workbook this way.
EXTERNAL_LINK_MARKER = "[1]"


def formula_text(value) -> str | None:
    """The formula string behind a cell, whether it's stored as a plain
    string or wrapped in an ArrayFormula."""
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def strip_external_formulas(worksheet) -> int:
    """Blanks every cell whose formula references the external workbook.
    Returns how many were cleared; formatting, merges and images are
    untouched.

    Note this does **not** catch local formulas that reference sheets the
    generated file won't have (SF2's `=HYPERLINK("#'PRINT CONTROL'!…")`
    is the example) — those have no `[1]` marker and each form clears its
    own.
    """
    cleared = 0
    for row in worksheet.iter_rows():
        for cell in row:
            text = formula_text(cell.value)
            if text and EXTERNAL_LINK_MARKER in text:
                cell.value = None
                cleared += 1
    return cleared


def assert_no_external_links(workbook) -> None:
    """Guard: a generated form must never still point at the school's
    master workbook. Raises rather than silently shipping a file that
    prompts to update links or shows stale numbers."""
    remaining = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                text = formula_text(cell.value)
                if text and EXTERNAL_LINK_MARKER in text:
                    remaining.append(f"{worksheet.title}!{cell.coordinate}")
    if remaining:
        raise AssertionError(
            f"{len(remaining)} external-link formula(s) survived, e.g. {remaining[:5]}"
        )
    if getattr(workbook, "_external_links", None):
        raise AssertionError("workbook still declares an external link book")


def anchor_map(worksheet) -> dict:
    """Maps every covered (row, col) to the top-left anchor of its merged
    range.

    Necessary because openpyxl raises on any write to a merged cell that
    isn't the anchor, and these templates' merge layouts **differ row by
    row** — a column that anchors a merge on one row can sit mid-merge on
    the next. Resolving per (row, col) is the only reliable way to write.
    """
    anchors: dict = {}
    for merged in worksheet.merged_cells.ranges:
        top_left = (merged.min_row, merged.min_col)
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                anchors[(row, col)] = top_left
    return anchors


def write(worksheet, anchors: dict, row: int, col: int, value) -> None:
    """Writes to (row, col), redirected to its merge anchor when needed."""
    target_row, target_col = anchors.get((row, col), (row, col))
    worksheet.cell(target_row, target_col).value = value


def write_ref(worksheet, anchors: dict, coordinate: str, value) -> None:
    cell = worksheet[coordinate]
    write(worksheet, anchors, cell.row, cell.column, value)


def clear_column(worksheet, column: int, last_row: int | None = None) -> None:
    """Blanks a whole helper column — the scaffolding these templates use
    to drive their own print macros, which has no meaning once the file
    leaves the master workbook."""
    for row in range(1, (last_row or worksheet.max_row) + 1):
        cell = worksheet.cell(row, column)
        if isinstance(cell, Cell):  # skip MergedCell, which is read-only
            cell.value = None


def replicate_images(base, copies) -> None:
    """Puts the header logos on every page of a multi-page form.

    Two openpyxl quirks make this less obvious than it looks:

    1. `Workbook.copy_worksheet` copies values, styles and merges but
       **silently drops images**, so page 2 prints without the DepEd and
       school seals while page 1 looks fine.
    2. A loaded image's bytes can only be read **once** — `_data()`
       consumes the underlying BytesIO and leaves it closed. Sharing one
       image object (or a shallow copy) across two sheets therefore saves
       the first and raises "I/O operation on closed file" on the second.

    So the bytes are captured once up front and a brand-new Image, with
    its own live buffer and anchor, is built for every sheet — including
    the base, whose original buffer the capture just spent.
    """
    if not base._images:
        return
    captured = [
        (image._data(), image.anchor, image.width, image.height) for image in base._images
    ]

    def rebuild():
        rebuilt = []
        for data, anchor, width, height in captured:
            image = XLImage(io.BytesIO(data))
            image.anchor = copy(anchor)
            image.width, image.height = width, height
            rebuilt.append(image)
        return rebuilt

    base._images = rebuild()
    for sheet in copies:
        sheet._images = rebuild()


def workbook_to_bytes(workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
