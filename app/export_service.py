"""Excel / CSV export (§52).

**LRN must survive as text.** §52 says so outright, and it's rule 10 in
CLAUDE.md: an LRN is a 12-digit identifier, not a quantity. Written as a
number, Excel renders `107041140016` in scientific notation and silently
eats a leading zero — so the column is written as a string *and* given an
explicit text number-format, and the CSV variant is quoted. There's a
test for the leading-zero case specifically.

Every export builds the same shape — a list of column names and a list of
row dicts — so adding one is a matter of writing the query, not the
plumbing.
"""

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models.academic_structure import GradeLevel, Section
from app.models.attendance import AttendanceRecord
from app.models.awards import AwardPolicy, AwardPolicyVersion, LearnerAward
from app.models.enums import ExportJobStatus
from app.models.admin import ExportJob
from app.models.grades import AnnualGradeSummary, TermGradeSummary
from app.models.learners import Enrollment, Learner
from app.models.organization import SchoolYear, Term
from app.report_card import build_learning_area_rows, load_report_context

# Columns whose values are identifiers rather than numbers. Excel would
# otherwise coerce them and lose leading zeros.
TEXT_COLUMNS = {"LRN"}

MAX_COLUMN_WIDTH = 42


@dataclass
class ExportTable:
    name: str
    columns: list[str]
    rows: list[dict]


def _dash(value, default="—"):
    return value if value is not None else default


def _grade(value):
    return int(value) if value is not None else None


def _learner_name(learner) -> str:
    return f"{learner.last_name}, {learner.first_name}" if learner else "?"


# --- The exports (§52) -----------------------------------------------------


def section_masterlist(session, section_id, school_year_id) -> ExportTable:
    enrollments, learners, _ = _roster(session, section_id, school_year_id)
    rows = []
    for index, enrollment in enumerate(enrollments, start=1):
        learner = learners.get(enrollment.learner_id)
        rows.append(
            {
                "No.": index,
                "LRN": learner.lrn or "" if learner else "",
                "Name": _learner_name(learner),
                "Sex": learner.sex.value if learner else "",
                "Birthdate": learner.birthdate.isoformat() if learner and learner.birthdate else "",
                "Enrollment Status": enrollment.enrollment_status.value,
            }
        )
    return ExportTable("Masterlist", ["No.", "LRN", "Name", "Sex", "Birthdate", "Enrollment Status"], rows)


def gradebook(session, section_id, school_year_id, term_number: int) -> ExportTable:
    """One row per learner, one column per subject, for a single term."""
    enrollments, learners, context = _roster(session, section_id, school_year_id)
    subject_names = _ordered_subjects(context, term_number)

    rows = []
    for enrollment in enrollments:
        row = {"LRN": _lrn(learners, enrollment), "Name": _learner_name(learners.get(enrollment.learner_id))}
        for subject_id, name in subject_names:
            offering_id = context.offerings_by_subject.get(subject_id, {}).get(term_number)
            row[name] = (
                _grade(context.term_grades.get((enrollment.id, offering_id)))
                if offering_id is not None
                else None
            )
        summary = (
            session.query(TermGradeSummary)
            .filter_by(enrollment_id=enrollment.id)
            .join(Term, Term.id == TermGradeSummary.term_id)
            .filter(Term.term_number == term_number)
            .one_or_none()
        )
        row["Term Average"] = _grade(summary.term_average) if summary else None
        rows.append(row)

    columns = ["LRN", "Name"] + [name for _, name in subject_names] + ["Term Average"]
    return ExportTable(f"Gradebook T{term_number}", columns, rows)


def final_grade_summary(session, section_id, school_year_id) -> ExportTable:
    """Final grades per learning area, following the §16 display rule so
    the export agrees with the report card."""
    enrollments, learners, context = _roster(session, section_id, school_year_id)
    summaries = {
        s.enrollment_id: s
        for s in session.query(AnnualGradeSummary)
        .filter(AnnualGradeSummary.enrollment_id.in_([e.id for e in enrollments]))
        .all()
    }

    area_names: list[str] = []
    per_learner = {}
    for enrollment in enrollments:
        rows = build_learning_area_rows(session, enrollment, context)
        per_learner[enrollment.id] = rows
        for row in rows:
            if row.name not in area_names:
                area_names.append(row.name)

    rows = []
    for enrollment in enrollments:
        row = {"LRN": _lrn(learners, enrollment), "Name": _learner_name(learners.get(enrollment.learner_id))}
        for area_row in per_learner[enrollment.id]:
            # A component's Final Grade is blank on the report card (§16);
            # the export matches, so the two can be reconciled.
            row[area_row.name] = None if area_row.is_component else _grade(area_row.final_grade)
        summary = summaries.get(enrollment.id)
        row["General Average"] = _grade(summary.general_average) if summary else None
        row["Completion"] = summary.completion_status.value if summary else "not computed"
        rows.append(row)

    return ExportTable(
        "Final Grades", ["LRN", "Name"] + area_names + ["General Average", "Completion"], rows
    )


def attendance_export(session, section_id, school_year_id, year: int, month: int) -> ExportTable:
    """Per-learner monthly totals, using the same engine the SF2 does."""
    from app.attendance_service import class_days_in_month, roster_for_month, summarize_month

    class_days = class_days_in_month(session, school_year_id, year, month)
    roster = roster_for_month(session, section_id, school_year_id, year, month)
    rows = []
    for enrollment, learner, window in roster:
        summary = summarize_month(session, enrollment, window, class_days)
        rows.append(
            {
                "LRN": learner.lrn or "",
                "Name": _learner_name(learner),
                "Sex": learner.sex.value,
                "Class Days": summary.eligible_days,
                "Present": summary.days_present,
                "Absent": summary.days_absent,
                "Late": summary.late_count,
                "Cutting": summary.cutting_count,
                "Not Encoded": summary.unencoded_days,
            }
        )
    return ExportTable(
        f"Attendance {year}-{month:02d}",
        ["LRN", "Name", "Sex", "Class Days", "Present", "Absent", "Late", "Cutting", "Not Encoded"],
        rows,
    )


def award_eligibility(session, section_id, school_year_id, award_policy_version_id) -> ExportTable:
    enrollments, learners, _ = _roster(session, section_id, school_year_id)
    version = session.get(AwardPolicyVersion, award_policy_version_id)
    policy = session.get(AwardPolicy, version.award_policy_id) if version else None
    terms = {t.id: t.name for t in session.query(Term).filter_by(school_year_id=school_year_id).all()}
    awards = (
        session.query(LearnerAward)
        .filter(
            LearnerAward.enrollment_id.in_([e.id for e in enrollments]),
            LearnerAward.award_policy_version_id == award_policy_version_id,
        )
        .all()
    )
    by_enrollment: dict = {}
    for award in awards:
        by_enrollment.setdefault(award.enrollment_id, []).append(award)

    rows = []
    for enrollment in enrollments:
        for award in by_enrollment.get(enrollment.id, [None]):
            learner = learners.get(enrollment.learner_id)
            rows.append(
                {
                    "LRN": learner.lrn or "" if learner else "",
                    "Name": _learner_name(learner),
                    "Policy": policy.name if policy else "",
                    "Term": terms.get(award.term_id, "Annual") if award else "",
                    "Result": award.award_result.value if award else "not computed",
                    "Award": (award.award_name or "") if award else "",
                    # §24: never a bare "Not Eligible" — the reason travels
                    # with the result, including into the export.
                    "Reason": award.reason if award else "",
                    "Overridden": "YES" if award and award.is_override else "",
                }
            )
    return ExportTable(
        "Award Eligibility",
        ["LRN", "Name", "Policy", "Term", "Result", "Award", "Reason", "Overridden"],
        rows,
    )


# --- Shared loading --------------------------------------------------------


def _roster(session, section_id, school_year_id):
    enrollments = (
        session.query(Enrollment)
        .filter_by(section_id=section_id, school_year_id=school_year_id)
        .join(Learner, Learner.id == Enrollment.learner_id)
        .order_by(Learner.last_name, Learner.first_name)
        .all()
    )
    learners = {
        l.id: l
        for l in session.query(Learner)
        .filter(Learner.id.in_([e.learner_id for e in enrollments]))
        .all()
    } if enrollments else {}
    context = load_report_context(session, enrollments)
    return enrollments, learners, context


def _lrn(learners, enrollment) -> str:
    learner = learners.get(enrollment.learner_id)
    return (learner.lrn or "") if learner else ""


def _ordered_subjects(context, term_number: int) -> list[tuple]:
    items = [
        (subject_id, context.subjects[subject_id].official_name)
        for subject_id, by_term in context.offerings_by_subject.items()
        if term_number in by_term and subject_id in context.subjects
    ]
    items.sort(key=lambda item: (context.subject_order.get(item[0], 9999), item[1]))
    return items


# --- Rendering -------------------------------------------------------------


def to_xlsx(table: ExportTable) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = table.name[:31]  # Excel's sheet-name limit

    worksheet.append(table.columns)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in table.rows:
        worksheet.append([row.get(column) for column in table.columns])

    for index, column in enumerate(table.columns, start=1):
        letter = get_column_letter(index)
        if column in TEXT_COLUMNS:
            # Both halves matter: the value is already a str, and the
            # format stops Excel re-interpreting it as a number when the
            # user edits or re-saves the file.
            for cell in worksheet[letter][1:]:
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)
        widest = max(
            [len(str(column))] + [len(str(row.get(column) or "")) for row in table.rows] or [0]
        )
        worksheet.column_dimensions[letter].width = min(widest + 2, MAX_COLUMN_WIDTH)

    worksheet.freeze_panes = "A2"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_csv(table: ExportTable) -> bytes:
    buffer = io.StringIO()
    # QUOTE_NONNUMERIC would coerce; quoting all keeps an LRN a quoted
    # string, which is the strongest hint a CSV can give a spreadsheet.
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL, lineterminator="\n")
    writer.writerow(table.columns)
    for row in table.rows:
        writer.writerow(["" if row.get(c) is None else row.get(c) for c in table.columns])
    return buffer.getvalue().encode("utf-8-sig")


def record_export(session, export_type: str, scope: dict, user_id, filename: str) -> ExportJob:
    job = ExportJob(
        export_type=export_type,
        requested_by_user_id=user_id,
        scope=scope,
        file_path=filename,
        status=ExportJobStatus.COMPLETE,
        completed_at=datetime.now(timezone.utc),
    )
    session.add(job)
    return job
