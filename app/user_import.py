"""Bulk user creation from a spreadsheet — the columns, the checks, and
the blank template.

Deliberately **not** an `ImportSpec` in `app/import_specs.py`, for two
reasons:

1. Every spec there commits inside the page's own transaction
   (`commit(session, parsed, user_id)`). Creating a user is half a remote
   call to Supabase Auth and half a database write, and the remote half
   cannot be rolled back — so it does not fit that contract and pretending
   it does would make a partial failure look atomic.
2. Registering it would need a new `import_job_type` enum value, which is
   an Alembic enum migration against a live database for a feature that
   the audit log already covers.

Everything here is **pure** — no session, no queries. The Users page
already loads the role list and the user list to draw itself, so it hands
those in and a file check costs zero extra round trips. That also lets
the rules be tested without a database.
"""

import io
import re
from dataclasses import dataclass

import openpyxl

from app.import_pipeline import ColumnSpec, RowError, ValidationResult
from app.naming import normalize_name

USER_COLUMNS = [
    ColumnSpec("email", "Email", True, ("emailaddress", "e-mail", "mail", "depedemail")),
    ColumnSpec(
        "full_name", "Full Name", True,
        ("name", "fullname", "teacher", "teachername", "employeename"),
    ),
    # Optional: a blank cell creates the account with no role, which is a
    # real state — the person can sign in and see nothing until an admin
    # grants one above. Better than refusing the whole file over it.
    ColumnSpec("roles", "Roles", False, ("role", "rolecode", "rolecodes", "access")),
]


@dataclass(frozen=True)
class _ColumnSet:
    """The one attribute `suggest_mapping` and `missing_required` read off
    an ImportSpec. Users aren't a spec (see the module docstring), but the
    header-matching is worth sharing rather than writing a second time."""

    columns: list[ColumnSpec]


USER_FILE = _ColumnSet(USER_COLUMNS)

# One cell may hold several roles. Teachers write the list every way a
# list gets written, so all of them are accepted.
_ROLE_SEPARATORS = re.compile(r"[,;/|\n]+")

# Deliberately loose. This is a typo guard — "juan.delacruz" with no
# domain, a stray space — not an attempt to decide what a deliverable
# address is, which no regular expression has ever managed.
_EMAIL = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")


def split_roles(raw) -> list[str]:
    """The role codes in one cell, upper-cased.

    Spaces become underscores so "subject teacher" matches
    SUBJECT_TEACHER — the codes are written for the database and nobody
    types them back with the underscore.
    """
    parts = _ROLE_SEPARATORS.split(str(raw or ""))
    return [p.strip().upper().replace(" ", "_") for p in parts if p.strip()]


def validate_users(rows: list[dict], *, role_codes, existing_emails) -> ValidationResult:
    """Checks a mapped file against the roles that exist and the accounts
    that already do.

    `role_codes` is every code in the `roles` table; `existing_emails` is
    every address in `users`, lower-cased. Both are passed in rather than
    queried, so validating re-runs for free while the admin looks at the
    preview.

    A row whose email already has an account is **not** an error and is
    **not** re-provisioned — it comes back with `exists=True` for the page
    to list as skipped. Re-uploading last month's file is the obvious
    mistake to make here, and `provision_user`'s ordinary behaviour for a
    known address is to reset its password; done in bulk that would hand
    every teacher in the file a password they don't know, mid-term, for no
    reason they'd understand. Resetting one password stays a deliberate
    single click on the user's own panel.
    """
    known = {str(code).upper() for code in role_codes}
    already = {str(email).strip().lower() for email in existing_emails}
    valid_list = ", ".join(sorted(known))

    result = ValidationResult()
    seen: dict[str, int] = {}

    for row in rows:
        number = row.get("__row__")
        errors_before = len(result.errors)

        email = str(row.get("email") or "").strip().lower()
        if not email:
            result.errors.append(RowError(number, "Email", "required"))
        elif not _EMAIL.fullmatch(email):
            result.errors.append(
                RowError(number, "Email", f"{email!r} doesn't look like an email address")
            )
        elif email in seen:
            result.errors.append(
                RowError(number, "Email", f"same address as row {seen[email]} in this file")
            )
        else:
            seen[email] = number

        full_name = normalize_name(row.get("full_name"))
        if not full_name:
            result.errors.append(RowError(number, "Full Name", "required"))

        codes = split_roles(row.get("roles"))
        unknown = [c for c in codes if c not in known]
        if unknown:
            result.errors.append(
                RowError(
                    number, "Roles",
                    f"unknown role {', '.join(unknown)} — use one of: {valid_list}",
                )
            )

        if len(result.errors) == errors_before:
            result.parsed.append(
                {
                    "__row__": number,
                    "email": email,
                    "full_name": full_name,
                    "role_codes": codes,
                    "exists": email in already,
                }
            )
    return result


def partition_existing(parsed: list[dict]) -> tuple[list[dict], list[dict]]:
    """(rows to create, rows already holding an account)."""
    return [r for r in parsed if not r["exists"]], [r for r in parsed if r["exists"]]


def template_bytes() -> bytes:
    """A blank .xlsx with just the header row.

    Headers only, on purpose: a template shipped with a worked example
    gets uploaded with the example still in it, and the school ends up
    with an account for a person who doesn't exist. The example lives on
    the page instead, where it can't be submitted.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Users"
    worksheet.append([column.label for column in USER_COLUMNS])
    for index, column in enumerate(USER_COLUMNS, start=1):
        worksheet.cell(row=1, column=index).font = openpyxl.styles.Font(bold=True)
        worksheet.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = 34
    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
