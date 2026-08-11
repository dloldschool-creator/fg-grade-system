"""Tests for app/sf2_report.py — the pure SF2 rules (§34), plus a guard
that the shipped template still matches the hardcoded cell layout."""

import io
import zipfile
from datetime import date

import openpyxl
import pytest

from app.excel_template import anchor_map, replicate_images, workbook_to_bytes
from app.models.enums import AttendanceStatus, EnrollmentStatus
from app.sf2_report import (
    COL_PRINT_CONTROL,
    COL_SUMMARY_F,
    COL_SUMMARY_M,
    DAY_COLS,
    FEMALE_FIRST_ROW,
    MALE_FIRST_ROW,
    ROWS_PER_SEX,
    ROW_DATE_SCAFFOLD,
    SHEET_NAME,
    TEMPLATE_PATH,
    _apply_print_setup,
    _fill_day_headers,
    _widen_summary_percentage_columns,
    first_friday_on_or_after,
    movement_remark,
    page_slice,
    paginate,
    printed_code,
)


# --- Pagination (§34) -----------------------------------------------------


@pytest.mark.parametrize(
    "males, females, expected",
    [
        (0, 0, 1),  # an empty section still prints one blank form
        (14, 11, 1),
        (25, 25, 1),  # exactly full
        (26, 25, 2),  # one male over
        (25, 26, 2),  # one female over
        (60, 10, 3),
    ],
)
def test_paginate_never_hides_learners(males, females, expected):
    """§34: overflow must produce additional pages, never drop learners.
    Male and female each have their own 25-row block, so the page count
    follows whichever sex overflows further — not the combined total. 30
    male + 30 female is 60 learners but still only 2 pages."""
    assert paginate(males, females) == expected


def test_paginate_is_driven_by_the_larger_sex_not_the_total():
    assert paginate(30, 30) == 2
    assert paginate(60, 0) == 3  # same 60 learners, all one sex, needs more pages


def test_page_slice_partitions_learners_without_loss():
    learners = list(range(60))
    pages = paginate(60, 0)
    collected = []
    for index in range(pages):
        collected.extend(page_slice(learners, index))
    assert collected == learners
    assert len(page_slice(learners, 0)) == ROWS_PER_SEX
    assert page_slice(learners, pages - 1) == learners[50:]
    assert page_slice(learners, pages) == []  # past the end


# --- Printed codes (§30) --------------------------------------------------


def test_present_prints_blank_and_absent_prints_x():
    """The paper form's convention is the inverse of the encoding UI's:
    present is an empty cell there, an explicit "P" on screen."""
    assert printed_code(AttendanceStatus.PRESENT) == ""
    assert printed_code(AttendanceStatus.ABSENT) == "X"
    assert printed_code(AttendanceStatus.LATE) == "T-L"
    assert printed_code(AttendanceStatus.CUTTING) == "T-C"


def test_unencoded_day_also_prints_blank():
    """Indistinguishable from present on paper — which is safe only
    because §33 blocks finalizing a month that still has un-encoded
    days."""
    assert printed_code(None) == ""


def test_movement_remark_names_the_movement_and_its_date():
    remark = movement_remark(EnrollmentStatus.TRANSFERRED_OUT, date(2026, 9, 12))
    assert remark == "Transferred Out 09/12/2026"


# --- Start-of-year reference date -----------------------------------------


def test_first_friday_is_anchored_to_the_school_year_start():
    """SY 2026-2027 opens Monday 8 June. June's calendar first Friday is
    the 5th — before anyone is enrolled — so the reference has to be the
    first Friday on or after opening, the 12th."""
    assert first_friday_on_or_after(date(2026, 6, 8)) == date(2026, 6, 12)


def test_first_friday_returns_the_start_date_when_it_is_itself_a_friday():
    assert first_friday_on_or_after(date(2026, 6, 12)) == date(2026, 6, 12)


# --- Template layout guard ------------------------------------------------


def test_template_still_matches_the_hardcoded_layout():
    """The cell coordinates in sf2_report are read off the shipped
    template. If the school swaps in a revised SF2, this fails loudly
    instead of silently writing learner names into the wrong cells."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    assert SHEET_NAME in workbook.sheetnames
    worksheet = workbook[SHEET_NAME]

    # The row labels that bound each 25-row block.
    assert str(worksheet.cell(43, 7).value).startswith("<=== MALE")
    assert str(worksheet.cell(69, 7).value).startswith("<=== FEMALE")
    assert worksheet.cell(70, 7).value == "Combined TOTAL Per Day"

    # Block sizes derived from those labels match §34's 25 rows.
    assert 43 - MALE_FIRST_ROW == ROWS_PER_SEX
    assert 69 - FEMALE_FIRST_ROW == ROWS_PER_SEX

    # Column headers.
    assert worksheet.cell(15, 1).value == "No."
    assert str(worksheet.cell(15, 7).value).startswith("NAME")
    assert worksheet.cell(17, 61).value == "ABSENT"
    assert worksheet.cell(17, 64).value == "PRESENT"
    assert str(worksheet.cell(15, 68).value).startswith("REMARKS")


def test_extra_pages_keep_the_header_logos():
    """`copy_worksheet` drops images, and a loaded image's bytes can only
    be read once — either bug alone makes page 2 print without the seals,
    or makes saving raise "I/O operation on closed file". This is the
    regression guard for both."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    base = workbook[SHEET_NAME]
    assert len(base._images) == 2, "template should carry the two header logos"

    extra = workbook.copy_worksheet(base)
    extra.title = "p2"
    assert extra._images == [], "copy_worksheet is expected to drop images"

    replicate_images(base, [extra])
    assert len(base._images) == 2
    assert len(extra._images) == 2

    # Saving must succeed — this is what failed when the sheets shared a
    # single already-consumed image buffer.
    saved = workbook_to_bytes(workbook)
    archive = zipfile.ZipFile(io.BytesIO(saved))
    drawings = [n for n in archive.namelist() if n.startswith("xl/drawings/drawing") and n.endswith(".xml")]
    assert len(drawings) == 2


def test_template_has_enough_day_columns_for_any_month():
    """The template's day slots have to cover the busiest month. SY
    2026-2027's is July at 23 class days; the form provides 25."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    # Every listed day column must carry a weekday formula/value in the
    # template, confirming the anchors are real day slots.
    for column in DAY_COLS:
        assert worksheet.cell(17, column).value is not None
    assert len(DAY_COLS) == 25
    assert len(DAY_COLS) >= 23


# --- Export presentation --------------------------------------------------


def test_print_control_column_exists_in_the_template():
    """The source workbook drove its own print macros from column CC
    ("Select Month in" plus a HYPERLINK to a 'PRINT CONTROL' sheet). The
    hyperlink is a *local* formula, so it survives external-link
    stripping and has to be cleared deliberately — this asserts the
    column is really the one being cleared."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    assert worksheet.cell(2, COL_PRINT_CONTROL).value == "Select Month in"
    assert "PRINT CONTROL" in str(worksheet.cell(3, COL_PRINT_CONTROL).value)


def test_print_setup_is_landscape_and_fits_the_width():
    """The template ships with no <pageSetup> at all, so without this the
    79-column form prints portrait at 100% across several pages.
    `fitToWidth` also does nothing unless the sheet's `fitToPage`
    property is set, which is easy to miss."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    _apply_print_setup(worksheet)

    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.fitToWidth == 1
    assert worksheet.page_setup.fitToHeight == 0  # may run to a second sheet
    assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True
    # Print area stops before the helper column.
    assert "$CA$112" in worksheet.print_area


def test_date_scaffold_row_is_cleared_not_populated():
    """Row 14 held date serials that the template's day-number and
    weekday rows read via formula. Those rows are written as plain values
    now, so nothing references row 14 — and since it sits above the table
    header, leaving dates there prints them over the form."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]

    # The template's own formulas confirm what the row was for.
    assert "$14" in str(worksheet.cell(16, DAY_COLS[0]).value)
    assert "$14" in str(worksheet.cell(17, DAY_COLS[0]).value)

    _fill_day_headers(worksheet, anchor_map(worksheet), [])
    for column in DAY_COLS:
        assert worksheet.cell(ROW_DATE_SCAFFOLD, column).value is None


def test_summary_percentage_columns_are_wide_enough():
    """Excel renders a *numeric* cell as ### when it's too narrow (text
    would just overflow into the neighbour). The template gives the M and
    F figures under 4 characters, which is fine for the counts but not
    for "100.00%" at seven — so both percentage rows printed as ###."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]

    def width(*columns):
        return sum(
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(c)].width or 8.43
            for c in columns
        )

    # The defect, as shipped.
    assert width(COL_SUMMARY_M, COL_SUMMARY_M + 1) < 7
    assert width(COL_SUMMARY_F) < 7

    _widen_summary_percentage_columns(worksheet)

    needed = len("100.00%")
    assert width(COL_SUMMARY_M, COL_SUMMARY_M + 1) >= needed
    assert width(COL_SUMMARY_F) >= needed
    # The wide Total column is left alone.
    assert worksheet.column_dimensions["BX"].width == 13.0
