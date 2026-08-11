"""Tests for the pure-Python xlsx -> PDF renderer (app/xlsx_render.py).

Each of these locks in a bug that actually happened while building it
against the real SF9 and SF2 templates. They're worth keeping because
every one of them produced output that looked *plausible* — a form that
is subtly wrong is worse than one that fails.
"""

import io

import openpyxl
import pytest

from app.excel_template import strip_external_formulas
from app.sf2_report import TEMPLATE_PATH as SF2_TEMPLATE
from app.sf2_report import _apply_print_setup as _apply_sf2_print_setup
from app.sf9_report import TEMPLATE_PATH as SF9_TEMPLATE
from app.sf9_report import _apply_print_setup as _apply_sf9_print_setup
from app.xlsx_render import (
    apply_number_format,
    column_dimension_map,
    column_width_to_points,
    page_count,
    plan_pages,
    resolve_font,
    sheet_geometry,
    workbook_to_pdf,
)


# --- Number formats --------------------------------------------------------


@pytest.mark.parametrize(
    "value, fmt, expected",
    [
        # `#` is an OPTIONAL digit and suppresses trailing zeros; `0` is
        # required. Treating them alike printed SF2's "No." column, which
        # uses `#.###`, as "1.000".
        (1, "#.###", "1"),
        (1.5, "#.###", "1.5"),
        (1, "0.00", "1.00"),
        # SF2 stores a percentage as a ratio; unformatted it prints "1"
        # where the division office expects "100.00%".
        (1.0, "0.00%", "100.00%"),
        (0.94, "0.00%", "94.00%"),
        (1234.5, "#,##0.00", "1,234.50"),
        (93.0, "General", "93"),
        (None, "General", ""),
        ("PASSED", "General", "PASSED"),
    ],
)
def test_number_formats(value, fmt, expected):
    assert apply_number_format(value, fmt) == expected


def test_an_unrecognised_format_falls_back_rather_than_guessing():
    """Excel's format language is far larger than the subset implemented;
    a wrong guess on an official form is worse than a plain number."""
    assert apply_number_format(42, "[Red][>100]0;;;") == "42"


# --- Geometry --------------------------------------------------------------


def test_grouped_column_definitions_apply_to_every_column_in_range():
    """A worksheet stores widths as `<col min= max= width=>` ranges keyed
    by the first letter only — SF9 has one entry under "P" governing
    columns 16-26. Reading them per letter left eleven columns on the
    default width, inflating the card from 12.25in to 16.11in wide and
    shrinking the whole form to 65% to make it fit."""
    workbook = openpyxl.load_workbook(SF9_TEMPLATE)
    worksheet = workbook["SF9"]
    dimensions = column_dimension_map(worksheet, worksheet.max_column)

    # P (16) through Z (26) all come from a single grouped definition.
    assert dimensions[16] is dimensions[26]
    assert {dimensions[col].width for col in range(16, 27)} == {dimensions[16].width}
    # H (8) through J (10) likewise — the three term columns.
    assert dimensions[8] is dimensions[10]


def test_sf9_geometry_matches_the_forms_real_size():
    """~9.5in tall is the figure the SF9 print setup was tuned against;
    the width must come out narrower than the height in landscape terms,
    or fit-to-page shrinks the card unnecessarily."""
    workbook = openpyxl.load_workbook(SF9_TEMPLATE)
    geometry = sheet_geometry(workbook["SF9"])
    assert 9.0 < geometry.height / 72 < 10.0
    assert 11.5 < geometry.width / 72 < 13.0


def test_a_hidden_column_takes_no_width():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.column_dimensions["B"].hidden = True
    geometry = sheet_geometry(worksheet, max_row=2, max_col=3)
    assert geometry.col_w[2] == 0.0


def test_column_width_conversion_uses_excels_own_padding():
    assert column_width_to_points(8.43) == pytest.approx(48.01, abs=0.05)


# --- Fonts -----------------------------------------------------------------


def test_unavailable_fonts_fall_back_to_metric_compatible_faces():
    """Arial isn't redistributable and Aparajita isn't on a Linux server.
    The old LibreOffice pipeline substituted these too — it only looked
    exact locally because Windows has them installed."""
    assert resolve_font("Arial", False, False) == "Helvetica"
    assert resolve_font("Aparajita", True, False) == "Helvetica-Bold"
    assert resolve_font("Carlito", False, True) == "Helvetica-Oblique"
    assert resolve_font(None, True, True) == "Helvetica-BoldOblique"


# --- End-to-end ------------------------------------------------------------


def _prepared(path, sheet):
    """A worksheet in the state the real pipeline renders it in.

    Both templates ship with **no `<pageSetup>` at all** — orientation,
    fit-to-page and margins are applied by the report module. Rendering the
    raw file measures a configuration that never reaches a printer.
    """
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook[sheet]
    strip_external_formulas(worksheet)
    if sheet == "SF9":
        _apply_sf9_print_setup(worksheet)
    else:
        _apply_sf2_print_setup(worksheet)
    return workbook, worksheet


def _render(path, sheet):
    workbook, _ = _prepared(path, sheet)
    return workbook_to_pdf(workbook)


@pytest.mark.parametrize("path, sheet", [(SF9_TEMPLATE, "SF9"), (SF2_TEMPLATE, "SF2")])
def test_templates_render_to_a_real_pdf(path, sheet):
    data = _render(path, sheet)
    assert data.startswith(b"%PDF-")
    assert len(data) > 5000  # a blank page would be far smaller


def _page_count(pdf: bytes) -> int:
    """Counts page objects without pulling in a PDF-reading dependency —
    `/Type /Pages` is the single page *tree* node, `/Type /Page` each leaf."""
    return pdf.count(b"/Type /Page\n") + pdf.count(b"/Type /Page ") + pdf.count(b"/Type /Page/")


def test_sf9_renders_landscape_on_a_single_page():
    """The Phase 10 print setup exists to keep the card on one sheet; a
    renderer that quietly spills onto a second page defeats it."""
    data = _render(SF9_TEMPLATE, "SF9")
    assert _page_count(data) == 1
    workbook, _ = _prepared(SF9_TEMPLATE, "SF9")
    assert page_count(workbook) == 1


def test_a_full_section_paginates_instead_of_shrinking_to_fit():
    """SF2 sets fitToWidth=1 with fitToHeight=0 — one page wide, as many
    pages tall as needed — and the form hides unused learner rows, so a
    3-learner section is short and a full one is 22.6in tall. Forcing that
    onto a single page needs 35% scale, which prints the form's 5pt text
    at 1.8pt: present on the page, and unreadable.
    """
    _, worksheet = _prepared(SF2_TEMPLATE, "SF2")
    for row in range(18, 68):  # every learner row the template allows
        worksheet.row_dimensions[row].hidden = False

    scale, bands = plan_pages(worksheet, 792, 612, sheet_geometry(worksheet))
    assert len(bands) > 1, "a full roster must paginate, not shrink"
    assert scale > 0.5, "scale must stay legible once it paginates"


def test_a_form_that_fits_exactly_does_not_spill_one_row():
    """SF9 pins both fitToWidth and fitToHeight to 1 and fills 575.9pt of
    its 576pt page. Without a break tolerance the last row lands on a
    second, nearly empty sheet."""
    _, worksheet = _prepared(SF9_TEMPLATE, "SF9")
    _, bands = plan_pages(worksheet, 792, 612, sheet_geometry(worksheet))
    assert len(bands) == 1


def test_a_component_row_keeps_its_indent_when_wrapped():
    """`report_card.COMPONENT_INDENT` is ten leading spaces, and that
    indent is what shows the Grade 11 language pair as two components of
    one parent learning area (§16). A plain split() drops it and the card
    reads as three unrelated subjects."""
    from app.report_card import COMPONENT_INDENT
    from app.xlsx_render import _wrap

    text = f"{COMPONENT_INDENT}Mabisang Komunikasyon"
    lines = _wrap(text, "Helvetica", 8, 200)
    assert lines[0].startswith(COMPONENT_INDENT)


def test_the_renderer_needs_no_external_program():
    """The whole point: no LibreOffice, so no 261 MB subprocess and no
    concurrency limit to police."""
    import app.xlsx_render as renderer

    source = open(renderer.__file__, encoding="utf-8").read()
    assert "subprocess" not in source
    assert "soffice" not in source
