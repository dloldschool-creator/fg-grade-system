"""Tests for the migration tooling (§51 import, §52 export).

The reading, mapping and value-parsing halves need no database. The
validators do, so those tests run against a real session and roll back.
"""

import io
import zipfile
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from app.database import SessionLocal
from app.export_service import ExportTable, to_xlsx
from app.import_pipeline import (
    apply_mapping,
    missing_required,
    parse_date,
    parse_grade,
    parse_lrn,
    read_table,
    suggest_mapping,
)
from app.import_specs import (
    LEARNER_IMPORT,
    TERM_GRADE_IMPORT,
    validate_learners,
    validate_term_grades,
)
from app.models.learners import Learner


@pytest.fixture
def session():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.rollback()
        session.close()


# --- Reading (§51 steps 1-2) -----------------------------------------------


def test_csv_is_read_with_row_numbers_matching_the_spreadsheet():
    """Errors are reported by row number, so it has to be the number the
    user sees — 1 is the header, so data starts at 2."""
    data = b"Last Name,First Name\nDela Cruz,Juan\nSantos,Maria\n"
    headers, rows = read_table(data, "masterlist.csv")
    assert headers == ["Last Name", "First Name"]
    assert [r["__row__"] for r in rows] == [2, 3]


def test_lrn_keeps_its_leading_zero_when_read():
    """Rule 10: an LRN is an identifier, not a number."""
    _, rows = read_table(b"LRN\n012345678901\n", "x.csv")
    assert rows[0]["LRN"] == "012345678901"


def test_excel_numeric_lrn_is_not_read_in_scientific_notation():
    """Excel hands a 12-digit LRN back as a float; a naive str() would
    render 1.07041140016e+11 and destroy the value."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["LRN"])
    worksheet.append([107041140016])
    buffer = io.BytesIO()
    workbook.save(buffer)

    _, rows = read_table(buffer.getvalue(), "x.xlsx")
    assert rows[0]["LRN"] == "107041140016"
    assert "e+" not in rows[0]["LRN"]


def test_blank_rows_are_skipped():
    _, rows = read_table(b"A,B\n1,2\n,\n3,4\n", "x.csv")
    assert len(rows) == 2


def test_excel_dates_come_through_as_iso_strings():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Birthdate"])
    worksheet.append([date(2009, 1, 15)])
    buffer = io.BytesIO()
    workbook.save(buffer)
    _, rows = read_table(buffer.getvalue(), "x.xlsx")
    assert rows[0]["Birthdate"] == "2009-01-15"


# --- Mapping (§51 step 3) --------------------------------------------------


def test_mapping_is_suggested_from_real_world_header_spellings():
    headers = ["Surname", "Given Name", "Gender", "Date of Birth", "Learner Reference Number"]
    mapping = suggest_mapping(headers, LEARNER_IMPORT)
    assert mapping["last_name"] == "Surname"
    assert mapping["first_name"] == "Given Name"
    assert mapping["sex"] == "Gender"
    assert mapping["birthdate"] == "Date of Birth"
    assert mapping["lrn"] == "Learner Reference Number"


def test_mapping_ignores_case_and_punctuation():
    mapping = suggest_mapping(["LAST_NAME", "first name"], LEARNER_IMPORT)
    assert mapping["last_name"] == "LAST_NAME"
    assert mapping["first_name"] == "first name"


def test_missing_required_columns_are_reported_by_label():
    absent = missing_required({"last_name": "Surname"}, LEARNER_IMPORT)
    assert "First Name" in absent
    assert "Last Name" not in absent
    assert "Middle Name" not in absent  # optional


def test_apply_mapping_preserves_the_row_number():
    rows = [{"Surname": "Dela Cruz", "__row__": 7}]
    mapped = apply_mapping(rows, {"last_name": "Surname"})
    assert mapped == [{"last_name": "Dela Cruz", "__row__": 7}]


# --- Value parsing ---------------------------------------------------------


@pytest.mark.parametrize("raw, expected", [("90", 90), ("90.4", 90), ("90.6", 91), ("", None)])
def test_grades_parse_and_round(raw, expected):
    value, error = parse_grade(raw)
    assert error is None
    assert value == expected


def test_a_blank_grade_is_not_a_zero():
    """Rule 2 at the import boundary: blank means not yet encoded."""
    assert parse_grade("") == (None, None)


@pytest.mark.parametrize("raw", ["abc", "101", "-5"])
def test_invalid_grades_are_reported(raw):
    value, error = parse_grade(raw)
    assert value is None and error


@pytest.mark.parametrize(
    "raw, expected",
    [("2009-01-15", date(2009, 1, 15)), ("15/01/2009", date(2009, 1, 15))],
)
def test_dates_parse_in_common_formats(raw, expected):
    assert parse_date(raw) == (expected, None)


@pytest.mark.parametrize("raw", ["2009-13-01", "not a date", "31/02/2009"])
def test_impossible_dates_are_reported(raw):
    """§51 names "impossible date" as a validation error."""
    value, error = parse_date(raw)
    assert value is None and error


@pytest.mark.parametrize("raw", ["12345", "abcdefghijkl", "1234567890123"])
def test_malformed_lrns_are_reported(raw):
    value, error = parse_lrn(raw)
    assert value is None and error


def test_valid_lrn_stays_a_string():
    value, error = parse_lrn("012345678901")
    assert error is None
    assert value == "012345678901"
    assert isinstance(value, str)


# --- Validation against the database (§51 steps 4-5) -----------------------


def _learner_rows(**overrides):
    row = {
        "__row__": 2,
        "last_name": "Testcase",
        "first_name": "Import",
        "middle_name": "",
        "extension_name": "",
        "sex": "MALE",
        "birthdate": "2009-01-15",
        "lrn": "",
        "section": "",
    }
    row.update(overrides)
    return [row]


def test_duplicate_lrn_against_the_database_is_rejected(session):
    """§51's first named error."""
    existing = session.query(Learner).filter(Learner.lrn.isnot(None)).first()
    if existing is None:
        pytest.skip("no learner with an LRN to collide with")
    result = validate_learners(session, _learner_rows(lrn=existing.lrn), {})
    assert not result.ok
    assert any("duplicate LRN" in e.message for e in result.errors)


def test_duplicate_lrn_within_the_same_file_is_rejected(session):
    rows = _learner_rows(lrn="099999999901")
    second = dict(rows[0])
    second["__row__"] = 3
    result = validate_learners(session, rows + [second], {})
    assert not result.ok
    assert any("row 2" in e.message for e in result.errors)


def test_a_valid_learner_row_parses_and_is_uppercased(session):
    result = validate_learners(session, _learner_rows(last_name="dela cruz"), {})
    assert result.ok, result.error_dicts()
    assert result.parsed[0]["last_name"] == "DELA CRUZ"


def test_learner_row_missing_required_fields_is_rejected(session):
    result = validate_learners(session, _learner_rows(last_name="", birthdate=""), {})
    assert not result.ok
    assert {e.column for e in result.errors} >= {"Last Name", "Birthdate"}


def _any_school_year_id(session):
    from app.models.organization import SchoolYear

    year = session.query(SchoolYear).first()
    if year is None:
        pytest.skip("no school year configured")
    return year.id


def test_unknown_section_and_subject_are_rejected(session):
    """Two more of §51's named errors."""
    rows = [
        {
            "__row__": 2,
            "lrn": "012345678901",
            "section": "NO SUCH SECTION",
            "subject": "NO SUCH SUBJECT",
            "term": "1",
            "grade": "90",
        }
    ]
    result = validate_term_grades(
        session, rows, {}, school_year_id=_any_school_year_id(session)
    )
    assert not result.ok
    messages = " ".join(e.message for e in result.errors)
    assert "unknown section" in messages
    assert "unknown subject" in messages


def test_term_grades_refuse_to_run_without_a_school_year(session):
    """A section name is unique only per grade level per school year, so
    without a year there is nothing to resolve it against. Matching on the
    name alone would silently write a grade into a same-named section of a
    different year — a row that looks fine and no report ever shows."""
    rows = [
        {
            "__row__": 2,
            "lrn": "012345678901",
            "section": "JOBS",
            "subject": "Whatever",
            "term": "1",
            "grade": "90",
        }
    ]
    result = validate_term_grades(session, rows, {}, school_year_id=None)
    assert not result.ok
    assert "school year" in result.errors[0].message
    assert not result.parsed


def test_term_grades_are_scoped_to_the_chosen_school_year(session):
    """A real section resolves in its own year and not in another one."""
    from app.models.academic_structure import Section
    from app.models.organization import SchoolYear

    section = session.query(Section).first()
    if section is None:
        pytest.skip("no sections configured")
    other_year = (
        session.query(SchoolYear).filter(SchoolYear.id != section.school_year_id).first()
    )
    if other_year is None:
        pytest.skip("only one school year configured")

    rows = [
        {
            "__row__": 2,
            "lrn": "012345678901",
            "section": section.name,
            "subject": "NO SUCH SUBJECT",
            "term": "1",
            "grade": "90",
        }
    ]
    in_its_year = validate_term_grades(
        session, rows, {}, school_year_id=section.school_year_id
    )
    in_other_year = validate_term_grades(session, rows, {}, school_year_id=other_year.id)

    assert not any("unknown section" in e.message for e in in_its_year.errors)
    assert any("unknown section" in e.message for e in in_other_year.errors)


def test_term_grades_reject_a_subject_not_offered_that_term(session):
    """The §51 error that matters most, because section_subject_offerings
    is the single source of truth for what a learner is graded on
    (rule 5). A grade for a term the subject doesn't run in must not be
    silently written."""
    from app.models.academic_structure import Section
    from app.models.organization import Term
    from app.models.subjects import SectionSubjectOffering, Subject

    from app.models.learners import Enrollment

    # Find a subject that genuinely doesn't run in every term — a
    # term-specific elective. Taking the first offering would often land
    # on a full-year subject and skip the very case being tested.
    terms = {t.id: t.term_number for t in session.query(Term).all()}
    by_pair: dict = {}
    for row in session.query(SectionSubjectOffering).all():
        term_number = terms.get(row.term_id)
        if term_number:
            by_pair.setdefault((row.section_id, row.subject_id), set()).add(term_number)

    candidate = next(
        ((pair, {1, 2, 3} - offered) for pair, offered in by_pair.items() if {1, 2, 3} - offered),
        None,
    )
    if candidate is None:
        pytest.skip("every configured subject runs in all three terms")
    (section_id, subject_id), unoffered = candidate
    section = session.get(Section, section_id)
    subject = session.get(Subject, subject_id)

    learner = (
        session.query(Learner)
        .join(Enrollment, Enrollment.learner_id == Learner.id)
        .filter(Enrollment.section_id == section_id, Learner.lrn.isnot(None))
        .first()
    )
    if learner is None:
        pytest.skip("no enrolled learner with an LRN in that section")

    rows = [
        {
            "__row__": 2,
            "lrn": learner.lrn,
            "section": section.name,
            "subject": subject.official_name,
            "term": str(sorted(unoffered)[0]),
            "grade": "90",
        }
    ]
    result = validate_term_grades(
        session, rows, {}, school_year_id=section.school_year_id
    )
    assert not result.ok
    assert any("is not offered" in e.message for e in result.errors)


# --- Enrolling from the learner import -------------------------------------


def test_a_blank_section_creates_the_learner_without_enrolling(session):
    """The column is optional, and leaving it blank must behave exactly as
    the import did before it existed."""
    result = validate_learners(session, _learner_rows(section=""), {}, school_year_id=None)
    assert result.ok, result.error_dicts()
    assert result.parsed[0]["section_id"] is None


def test_an_unknown_section_is_rejected(session):
    from app.models.organization import SchoolYear

    school_year = session.query(SchoolYear).first()
    if school_year is None:
        pytest.skip("no school year")
    result = validate_learners(
        session, _learner_rows(section="NO SUCH SECTION"), {}, school_year_id=school_year.id
    )
    assert not result.ok
    assert any("unknown section" in e.message for e in result.errors)


def test_a_section_without_a_school_year_is_rejected(session):
    """The school year is chosen on the page, not repeated on 1,200 rows.
    A file naming a section with no year selected cannot be resolved, and
    guessing which year to enrol into would be worse than refusing."""
    result = validate_learners(
        session, _learner_rows(section="ANY"), {}, school_year_id=None
    )
    assert not result.ok
    assert any("school year" in e.message for e in result.errors)


def test_a_real_section_carries_its_grade_level(session):
    """The grade level comes from the section rather than the file, so the
    two can never disagree."""
    from app.models.academic_structure import Section
    from app.models.organization import SchoolYear

    school_year = session.query(SchoolYear).first()
    if school_year is None:
        pytest.skip("no school year")
    section = session.query(Section).filter_by(school_year_id=school_year.id).first()
    if section is None:
        pytest.skip("no sections")

    result = validate_learners(
        session, _learner_rows(section=section.name), {}, school_year_id=school_year.id
    )
    assert result.ok, result.error_dicts()
    row = result.parsed[0]
    assert row["section_id"] == section.id
    assert row["grade_level_id"] == section.grade_level_id
    assert row["school_year_id"] == school_year.id


# --- An adviser enrolling their own class (§3C) ----------------------------


def _two_sections_in_one_year(session):
    """Two sections of the same school year, with an adviser put on the
    first one *inside this test's transaction*.

    Nothing is committed — the fixture rolls back — and a real user id is
    used so the FK still holds when the validator's own query autoflushes
    the change.

    **The tests below pass `str(user.id)`, not `user.id`.** That is what
    the app passes — `AuthUser.id` is our `users.id` as a str — and
    calling these with the ORM's `uuid.UUID` is how the first version of
    this feature shipped a comparison that refused every adviser their
    own section while every test passed.
    """
    from app.models.academic_structure import Section
    from app.models.organization import SchoolYear
    from app.models.rbac import User

    user = session.query(User).first()
    if user is None:
        pytest.skip("no users configured")
    for school_year in session.query(SchoolYear).all():
        sections = (
            session.query(Section)
            .filter_by(school_year_id=school_year.id)
            .order_by(Section.name)
            .all()
        )
        # A name shared by two grade levels is a case of its own (below):
        # it is ambiguous for a Registrar and resolves to the adviser's
        # own for an adviser, so neither pick here may be one of those.
        by_name: dict = {}
        for section in sections:
            by_name.setdefault(section.name, []).append(section)
        unique = [group[0] for group in by_name.values() if len(group) == 1]
        if len(unique) >= 2:
            mine, theirs = unique[:2]
            mine.adviser_user_id = user.id
            if theirs.adviser_user_id == user.id:
                theirs.adviser_user_id = None
            return school_year, mine, theirs, user
    pytest.skip("no school year with two differently-named sections")


def test_an_adviser_may_enrol_into_a_section_they_advise(session):
    school_year, mine, _theirs, user = _two_sections_in_one_year(session)
    result = validate_learners(
        session,
        _learner_rows(section=mine.name),
        {},
        school_year_id=school_year.id,
        adviser_user_id=str(user.id),
    )
    assert result.ok, result.error_dicts()
    assert result.parsed[0]["section_id"] == mine.id
    assert result.parsed[0]["grade_level_id"] == mine.grade_level_id


def test_the_adviser_id_is_matched_whether_it_is_a_str_or_a_uuid(session):
    """The regression this feature shipped with. The page hands over
    `AuthUser.id`, a **str**; the column holds a `uuid.UUID`. Postgres
    coerces one to the other — so the panel listed the adviser's sections
    correctly — while Python's `==` quietly said no to every one of them,
    and the adviser saw "MUSK is not one of your sections" for the
    section she advises."""
    school_year, mine, _theirs, user = _two_sections_in_one_year(session)
    for adviser_user_id in (str(user.id), user.id):
        result = validate_learners(
            session,
            _learner_rows(section=mine.name),
            {},
            school_year_id=school_year.id,
            adviser_user_id=adviser_user_id,
        )
        assert result.ok, (type(adviser_user_id).__name__, result.error_dicts())
        assert result.parsed[0]["section_id"] == mine.id


def test_an_adviser_is_refused_a_section_they_do_not_advise(session):
    """§3C: an adviser sees their own sections. The row is refused rather
    than the column dropped, so the file says which learners were left
    out instead of quietly creating all of them unenrolled."""
    school_year, _mine, theirs, user = _two_sections_in_one_year(session)
    result = validate_learners(
        session,
        _learner_rows(section=theirs.name),
        {},
        school_year_id=school_year.id,
        adviser_user_id=str(user.id),
    )
    assert not result.parsed
    assert any("not one of your sections" in e.message for e in result.errors)


def test_a_registrar_is_not_scoped_to_any_adviser(session):
    """The same file with no adviser id resolves the other section fine —
    the scoping is the caller's choice, not a new rule for everyone."""
    school_year, _mine, theirs, _user = _two_sections_in_one_year(session)
    result = validate_learners(
        session, _learner_rows(section=theirs.name), {}, school_year_id=school_year.id
    )
    assert result.ok, result.error_dicts()
    assert result.parsed[0]["section_id"] == theirs.id


def test_a_name_in_two_grade_levels_resolves_to_the_adviser_s_own(session):
    """A section name is unique only per grade level, so the same name can
    exist in Grade 11 and Grade 12. That is ambiguous for a Registrar and
    is refused — but an adviser holding exactly one of the two has named
    it unambiguously."""
    from app.import_specs import _section_lookup
    from app.models.academic_structure import Section
    from app.models.organization import SchoolYear
    from app.models.rbac import User

    user = session.query(User).first()
    if user is None:
        pytest.skip("no users configured")
    pair = None
    for school_year in session.query(SchoolYear).all():
        sections = session.query(Section).filter_by(school_year_id=school_year.id).all()
        by_grade = {}
        for section in sections:
            by_grade.setdefault(section.grade_level_id, []).append(section)
        if len(by_grade) < 2:
            continue
        (_g1, first), (_g2, second) = list(by_grade.items())[:2]
        pair = (school_year, first[0], second[0])
        break
    if pair is None:
        pytest.skip("no school year with sections in two grade levels")

    school_year, mine, clash = pair
    # Give them the same name and only one of them to the adviser.
    clash.name = mine.name
    mine.adviser_user_id = user.id
    if clash.adviser_user_id == user.id:
        clash.adviser_user_id = None
    key = mine.name.strip().upper()

    _by_name, ambiguous = _section_lookup(session, school_year.id)
    assert key in ambiguous

    by_name, ambiguous = _section_lookup(session, school_year.id, str(user.id))
    assert key not in ambiguous
    assert by_name[key].id == mine.id


def test_the_section_column_is_optional_and_auto_detected():
    from app.import_pipeline import suggest_mapping

    column = next(c for c in LEARNER_IMPORT.columns if c.field == "section")
    assert not column.required
    for header in ("Section", "Section Name", "Class"):
        assert suggest_mapping([header], LEARNER_IMPORT).get("section") == header


# --- Export (§52) ----------------------------------------------------------


def _table() -> ExportTable:
    return ExportTable(
        "Masterlist",
        ["LRN", "Name", "Grade"],
        [
            {"LRN": "012345678901", "Name": "DELA CRUZ, JUAN", "Grade": 90},
            {"LRN": "107041140016", "Name": "SANTOS, MARIA", "Grade": None},
        ],
    )


def test_xlsx_export_writes_lrn_as_text_with_its_leading_zero():
    """§52: "Exports should preserve LRN as text." Written as a number,
    Excel shows 1.2345678901E+11 and drops the leading zero."""
    workbook = openpyxl.load_workbook(io.BytesIO(to_xlsx(_table())))
    worksheet = workbook.active
    cell = worksheet["A2"]
    assert cell.value == "012345678901"
    assert isinstance(cell.value, str)
    assert cell.number_format == "@"  # text, so Excel won't re-coerce on save


def test_xlsx_export_is_a_real_workbook_with_a_header_row():
    data = to_xlsx(_table())
    assert zipfile.ZipFile(io.BytesIO(data)).namelist()  # valid xlsx container
    worksheet = openpyxl.load_workbook(io.BytesIO(data)).active
    assert [c.value for c in worksheet[1]] == ["LRN", "Name", "Grade"]
    assert worksheet.freeze_panes == "A2"


def test_xlsx_export_writes_a_missing_grade_as_blank_not_zero():
    """Rule 2 at the export boundary: the two CSV tests that used to cover
    this went with to_csv(), and .xlsx is now the only format."""
    worksheet = openpyxl.load_workbook(io.BytesIO(to_xlsx(_table()))).active
    grade_column = [c.value for c in worksheet["C"]]
    assert grade_column[0] == "Grade"
    assert None in grade_column[1:]
    assert 0 not in grade_column[1:]
