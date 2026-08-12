"""SF4-SHS — Monthly Learners' Movement and Attendance.

The counting rules are tested without a database: `aggregate` takes plain
collections precisely so the arithmetic can be pinned independently of
how the rows are fetched.
"""

from dataclasses import dataclass
from datetime import date

import openpyxl
import pytest

from app.models.enums import AttendanceStatus, EnrollmentStatus, Sex
from app.sf4_report import (
    GRADE_BLOCKS,
    MOVEMENT_COLUMNS,
    ROW_GRAND_TOTAL,
    SHEET_NAME,
    TEMPLATE_PATH,
    Sf4Row,
    Tally,
    aggregate,
    month_bounds,
    term_for_month,
)


# --- Stand-ins for the ORM rows -------------------------------------------


@dataclass
class FakeDay:
    id: int
    calendar_date: date


@dataclass
class FakeRecord:
    status: AttendanceStatus


@dataclass
class FakeMovement:
    enrollment_id: int
    movement_type: EnrollmentStatus
    effective_date: date


@dataclass
class FakeEnrollment:
    id: int
    learner_id: int
    section_id: int = 1
    grade_level_id: int = 11


@dataclass
class FakeLearner:
    id: int
    sex: Sex


@dataclass
class Named:
    name: str
    code: str = ""


@dataclass
class FakeSection:
    id: int = 1
    grade_level_id: int = 11
    track_id: int = 1
    strand_id: int = 1


JUNE = [FakeDay(i, date(2026, 6, day)) for i, day in enumerate([8, 9, 10, 11, 12], start=1)]
YEAR_START = date(2026, 6, 8)

CONTEXT = dict(
    sections={1: FakeSection()},
    tracks={1: Named("Academic")},
    strands={1: Named("STEM")},
    grade_levels={11: Named("Grade 11", "G11"), 12: Named("Grade 12", "G12")},
    class_days=JUNE,
    school_year_start=YEAR_START,
    year=2026,
    month=6,
)


def _run(enrollments, learners, movements=(), attendance=None):
    return aggregate(
        enrollments=enrollments,
        learners=learners,
        movements=list(movements),
        attendance=attendance or {},
        **CONTEXT,
    )


# --- Counting --------------------------------------------------------------


def test_registered_counts_learners_still_on_the_roll_at_month_end():
    rows = _run(
        [FakeEnrollment(1, 1), FakeEnrollment(2, 2)],
        {1: FakeLearner(1, Sex.MALE), 2: FakeLearner(2, Sex.FEMALE)},
    )
    assert rows[0].registered.male == 1
    assert rows[0].registered.female == 1
    assert rows[0].registered.total == 2


def test_a_learner_who_transferred_out_mid_month_is_not_registered_at_month_end():
    """They still appear on that month's SF2 with a remark (§32), but SF4
    asks who is on the roll at the *end* of the month."""
    rows = _run(
        [FakeEnrollment(1, 1)],
        {1: FakeLearner(1, Sex.MALE)},
        movements=[FakeMovement(1, EnrollmentStatus.TRANSFERRED_OUT, date(2026, 6, 10))],
    )
    assert rows[0].registered.total == 0


def test_movements_split_into_before_and_during_the_month():
    """The form wants (A) cumulative to the end of last month and (B) this
    month separately, then their sum."""
    rows = _run(
        [FakeEnrollment(1, 1), FakeEnrollment(2, 2)],
        {1: FakeLearner(1, Sex.MALE), 2: FakeLearner(2, Sex.MALE)},
        movements=[
            FakeMovement(1, EnrollmentStatus.DROPPED, date(2026, 5, 20)),
            FakeMovement(2, EnrollmentStatus.DROPPED, date(2026, 6, 11)),
        ],
    )
    before, during = rows[0].movement(EnrollmentStatus.DROPPED)
    assert before.total == 1
    assert during.total == 1


def test_late_and_cutting_count_as_present():
    """§30 — the learner was in school. Counting them absent here would
    contradict SF2 for the same month."""
    enrollments = [FakeEnrollment(1, 1)]
    learners = {1: FakeLearner(1, Sex.MALE)}
    attendance = {
        (1, JUNE[0].id): FakeRecord(AttendanceStatus.LATE),
        (1, JUNE[1].id): FakeRecord(AttendanceStatus.CUTTING),
        (1, JUNE[2].id): FakeRecord(AttendanceStatus.PRESENT),
        (1, JUNE[3].id): FakeRecord(AttendanceStatus.ABSENT),
    }
    rows = _run(enrollments, learners, attendance=attendance)
    assert rows[0].present_days.total == 3


def test_attendance_is_measured_against_eligible_days_only():
    """A learner who joined late could not attend the earlier days (§31);
    counting those against them understates the school."""
    enrollments = [FakeEnrollment(1, 1)]
    learners = {1: FakeLearner(1, Sex.MALE)}
    rows = _run(
        enrollments, learners,
        movements=[FakeMovement(1, EnrollmentStatus.LATE_ENROLLMENT, date(2026, 6, 11))],
        attendance={(1, JUNE[3].id): FakeRecord(AttendanceStatus.PRESENT)},
    )
    assert rows[0].eligible_days.total == 2, "only the 11th and 12th were attendable"
    assert rows[0].percentage().total == pytest.approx(50.0)


def test_two_strands_sharing_a_name_stay_separate_rows():
    """Rows are keyed on ids, never on names.

    The school briefly had two TechPro strands both called "ICT Support
    and Computer Programming Technologies" under different codes. Grouping
    by name merged them into a single row, quietly under-reporting the
    track/strand breakdown on a form filed with the division — the kind of
    wrong that looks like a correct form.
    """
    context = dict(CONTEXT)
    context["sections"] = {
        1: FakeSection(id=1, strand_id=1),
        2: FakeSection(id=2, strand_id=2),
    }
    context["strands"] = {
        1: Named("ICT Support and Computer Programming Technologies"),
        2: Named("ICT Support and Computer Programming Technologies"),
    }
    rows = aggregate(
        enrollments=[
            FakeEnrollment(1, 1, section_id=1),
            FakeEnrollment(2, 2, section_id=2),
        ],
        learners={1: FakeLearner(1, Sex.MALE), 2: FakeLearner(2, Sex.FEMALE)},
        movements=[],
        attendance={},
        **context,
    )
    assert len(rows) == 2, "same-named strands must not collapse into one row"
    assert {r.registered.total for r in rows} == {1}


def test_the_same_strand_across_two_sections_is_one_row():
    """The other half of the rule: SF4 reports per Track/Strand, not per
    section, so two sections of the same strand combine."""
    context = dict(CONTEXT)
    context["sections"] = {
        1: FakeSection(id=1, strand_id=1),
        2: FakeSection(id=2, strand_id=1),
    }
    rows = aggregate(
        enrollments=[
            FakeEnrollment(1, 1, section_id=1),
            FakeEnrollment(2, 2, section_id=2),
        ],
        learners={1: FakeLearner(1, Sex.MALE), 2: FakeLearner(2, Sex.MALE)},
        movements=[],
        attendance={},
        **context,
    )
    assert len(rows) == 1
    assert rows[0].registered.total == 2


# --- The percentage trap ---------------------------------------------------


def test_a_percentage_total_is_recomputed_not_summed():
    """62.5% of males plus 100% of females is not 162.5% of learners. SF2
    reported 200% from exactly this."""
    entry = Sf4Row(11, "Academic", "STEM")
    entry.present_days = Tally(male=30, female=16)
    entry.eligible_days = Tally(male=48, female=16)

    percentage = entry.percentage()
    assert percentage.male == pytest.approx(62.5)
    assert percentage.female == pytest.approx(100.0)
    assert percentage.total == pytest.approx(71.875)
    assert percentage.total < 100


def test_counts_and_daily_averages_still_sum():
    """The override is for percentages only — a headcount total really is
    male plus female."""
    assert Tally(male=3, female=1).total == 4
    entry = Sf4Row(11, "A", "B")
    entry.present_days = Tally(male=30, female=16)
    average = entry.daily_average(16)
    assert average.total == pytest.approx(average.male + average.female)


# --- Layout ----------------------------------------------------------------


def test_the_template_still_has_the_rows_this_module_writes_to():
    """Guards against a revised template silently shifting the grid."""
    worksheet = openpyxl.load_workbook(TEMPLATE_PATH)[SHEET_NAME]
    assert worksheet.cell(23, 1).value == "TOTAL FOR GRADE 11"
    assert worksheet.cell(35, 1).value == "TOTAL FOR GRADE 12"
    assert worksheet.cell(ROW_GRAND_TOTAL, 1).value == "GRAND TOTAL"
    # Row 11 labels every column as M / F / T.
    for first_col in [3, 6, 9] + list(MOVEMENT_COLUMNS.values()):
        assert [worksheet.cell(11, first_col + i).value for i in range(3)] == ["M", "F", "T"]


def test_the_grade_blocks_stop_before_their_total_rows():
    for _grade, (first_row, last_row, total_row) in GRADE_BLOCKS.items():
        assert first_row <= last_row < total_row


def test_month_bounds_cover_the_whole_month():
    assert month_bounds(2026, 6) == (date(2026, 6, 1), date(2026, 6, 30))
    assert month_bounds(2027, 2) == (date(2027, 2, 1), date(2027, 2, 28))


# --- The semester field ----------------------------------------------------


@dataclass
class FakeTerm:
    name: str
    start_date: date
    end_date: date


TERMS = [
    FakeTerm("Term 1", date(2026, 6, 8), date(2026, 9, 15)),
    FakeTerm("Term 2", date(2026, 9, 16), date(2026, 12, 18)),
    FakeTerm("Term 3", date(2027, 1, 4), date(2027, 4, 8)),
]


def test_the_semester_field_names_the_term_the_month_falls_in():
    """SF4 is a monthly report, so it needs no term-to-semester mapping —
    which is why it can be built while SF5 waits for the division. The
    form is not reinterpreted; the school's real period is named."""
    assert term_for_month(TERMS, 2026, 6).name == "Term 1"
    assert term_for_month(TERMS, 2026, 10).name == "Term 2"
    assert term_for_month(TERMS, 2027, 2).name == "Term 3"


def test_a_month_spanning_two_terms_takes_the_earlier_one():
    """September holds the Term 1/Term 2 boundary."""
    assert term_for_month(TERMS, 2026, 9).name == "Term 1"


def test_a_month_outside_every_term_has_no_semester():
    assert term_for_month(TERMS, 2026, 5) is None
