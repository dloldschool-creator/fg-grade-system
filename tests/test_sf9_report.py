"""Tests for app/sf9_report.py and app/report_card.py — the §16
combined-language display rule as it reaches the printed card, plus a
guard that the shipped template still matches the hardcoded layout."""

import io
import re
import zipfile
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from app.excel_template import anchor_map, strip_external_formulas, workbook_to_bytes
from app.report_card import COMPONENT_INDENT, LearningAreaRow
from app.sf9_report import (
    ATTENDANCE_MONTHS,
    COL_FINAL_GRADE,
    COL_LEARNING_AREA,
    COL_REMARKS,
    COL_TERM,
    COL_TERM_OFFERED_FLAGS,
    LEARNING_AREA_FIRST_ROW,
    LEARNING_AREA_LAST_ROW,
    MAX_LEARNING_AREAS,
    MONTH_ABBREVIATIONS,
    SHEET_NAME,
    TEMPLATE_PATH,
    _age_on,
    _apply_print_setup,
    _fill_learning_areas,
    _term_flags,
    _grade_level_number,
    next_level_after,
)

D = Decimal


# --- The §16 rule, as printed ---------------------------------------------


def _row(name, terms, final, remark, component=False) -> LearningAreaRow:
    return LearningAreaRow(
        name=name, term_grades=terms, final_grade=final, remark=remark, is_component=component
    )


def _filled_sheet(rows, exit_line=None):
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    strip_external_formulas(worksheet)
    anchors = anchor_map(worksheet)
    _fill_learning_areas(worksheet, anchors, rows, exit_line)

    def read(offset, column):
        row = LEARNING_AREA_FIRST_ROW + offset
        target = anchors.get((row, column), (row, column))
        return worksheet.cell(*target).value

    read.worksheet = worksheet
    return read


def test_combined_parent_carries_the_final_grade_and_components_stay_blank():
    """§16, on the printed card: the parent row shows the combined term
    grades AND a Final Grade; each component shows its own term grades but
    its Final Grade cell is blank. This is the rule CLAUDE.md calls the
    biggest source of General Average bugs, so it's asserted on the
    generated cells, not just on the row objects."""
    read = _filled_sheet(
        [
            _row("Effective Communication / Mabisang Komunikasyon",
                 {1: D(91), 2: D(94), 3: D(93)}, D(93), "PASSED"),
            _row("Effective Communication", {1: D(90), 2: D(98), 3: D(92)}, None, None, True),
            _row("Mabisang Komunikasyon", {1: D(92), 2: D(90), 3: D(93)}, None, None, True),
        ]
    )

    # Parent.
    assert read(0, COL_LEARNING_AREA) == "Effective Communication / Mabisang Komunikasyon"
    assert [read(0, COL_TERM[t]) for t in (1, 2, 3)] == [91, 94, 93]
    assert read(0, COL_FINAL_GRADE) == 93
    assert read(0, COL_REMARKS) == "PASSED"

    # Components: own term grades, blank Final Grade and Remarks.
    for offset in (1, 2):
        assert read(offset, COL_FINAL_GRADE) is None
        assert read(offset, COL_REMARKS) is None
    assert [read(1, COL_TERM[t]) for t in (1, 2, 3)] == [90, 98, 92]
    assert [read(2, COL_TERM[t]) for t in (1, 2, 3)] == [92, 90, 93]


def test_component_rows_are_indented_under_their_parent():
    read = _filled_sheet(
        [
            _row("Parent Area", {1: D(90)}, D(90), "PASSED"),
            _row("A Component", {1: D(90)}, None, None, True),
        ]
    )
    assert read(1, COL_LEARNING_AREA) == f"{COMPONENT_INDENT}A Component"
    assert read(0, COL_LEARNING_AREA) == "Parent Area"


def test_unused_rows_are_left_blank_not_filled_with_placeholders():
    """§35: "Do not print unused placeholder subjects. Blank unused rows
    may remain visually blank to preserve the official template.\""""
    read = _filled_sheet([_row("Only Subject", {1: D(88)}, D(88), "PASSED")])
    for offset in range(1, MAX_LEARNING_AREAS):
        assert read(offset, COL_LEARNING_AREA) is None
        assert read(offset, COL_FINAL_GRADE) is None
        assert read(offset, COL_REMARKS) is None


def test_grades_print_as_whole_numbers():
    """Stored Numeric(5,2) but integral by construction (§60) — the card
    must not show "93.00"."""
    read = _filled_sheet([_row("Subject", {1: D("93.00")}, D("93.00"), "PASSED")])
    assert read(0, COL_TERM[1]) == 93
    assert read(0, COL_FINAL_GRADE) == 93
    assert not isinstance(read(0, COL_FINAL_GRADE), Decimal)


def test_a_missing_term_grade_prints_blank_not_zero():
    """The NULL-is-not-zero rule reaching the form: a term the learner
    wasn't offered the subject in stays empty."""
    read = _filled_sheet([_row("One-Term Elective", {1: D(94)}, D(94), "PASSED")])
    assert read(0, COL_TERM[1]) == 94
    assert read(0, COL_TERM[2]) is None
    assert read(0, COL_TERM[3]) is None


# --- Exit status (§35 amendment, 2026-09-05) --------------------------------


def test_exit_line_replaces_every_rows_own_remark_with_one_merged_cell():
    """A learner who dropped out mid-year has every remaining subject's
    Final Grade genuinely None — printing INCOMPLETE on each of them would
    be true but not what happened. The Remarks column instead collapses
    into one cell naming the actual status."""
    read = _filled_sheet(
        [
            _row("Subject A", {1: D(88)}, D(88), "PASSED"),
            _row("Subject B", {1: D(75)}, D(75), "PASSED"),
            _row("Subject C", {1: None}, None, "INCOMPLETE"),
        ],
        exit_line="Dropped as of 08/30/2026 due to Child labor, work",
    )
    assert read(0, COL_REMARKS) == "Dropped as of 08/30/2026 due to Child labor, work"
    # Each row's own PASSED/PASSED/INCOMPLETE never gets written — the
    # merge is the only remark and it belongs to the anchor cell alone.
    assert read(1, COL_REMARKS) is None
    assert read(2, COL_REMARKS) is None


def test_exit_line_merge_spans_exactly_the_printed_rows():
    read = _filled_sheet(
        [
            _row("Subject A", {1: D(88)}, D(88), "PASSED"),
            _row("Subject B", {1: D(75)}, D(75), "PASSED"),
        ],
        exit_line="Dropped as of 08/30/2026 due to Illness",
    )
    ranges = [
        m for m in read.worksheet.merged_cells.ranges
        if m.min_col == COL_REMARKS and m.min_row == LEARNING_AREA_FIRST_ROW
    ]
    assert len(ranges) == 1
    assert ranges[0].max_row == LEARNING_AREA_FIRST_ROW + 1  # exactly 2 rows, not the full block
    assert ranges[0].max_col == COL_REMARKS + 1  # the template's own L:M width


def test_exit_line_with_a_single_printed_row_uses_the_templates_own_merge():
    """One subject means the template's per-row L:M merge already is the
    cell wanted — nothing to unmerge or rebuild."""
    read = _filled_sheet(
        [_row("Only Subject", {1: D(88)}, D(88), "PASSED")],
        exit_line="NLS as of 08/30/2026",
    )
    assert read(0, COL_REMARKS) == "NLS as of 08/30/2026"
    ranges = [
        m for m in read.worksheet.merged_cells.ranges
        if m.min_col == COL_REMARKS and m.min_row == LEARNING_AREA_FIRST_ROW
    ]
    assert len(ranges) == 1
    assert ranges[0].max_row == LEARNING_AREA_FIRST_ROW


def test_exit_line_does_not_touch_the_general_average_remark_row():
    """§35 amendment: the merge stops at the last printed subject row — row
    32 (General Average) keeps its own separate Remarks cell, computed
    elsewhere and untouched by this."""
    read = _filled_sheet(
        [_row("Subject A", {1: D(88)}, D(88), "PASSED")],
        exit_line="Dropped as of 08/30/2026 due to Child labor, work",
    )
    ga_ranges = [m for m in read.worksheet.merged_cells.ranges if m.min_row == 32 and m.min_col == COL_REMARKS]
    assert len(ga_ranges) == 1
    assert ga_ranges[0].max_row == 32, "the General Average remark merge must not be swallowed"


def test_no_exit_line_leaves_the_templates_own_per_row_merges_alone():
    read = _filled_sheet([_row("Subject A", {1: D(88)}, D(88), "PASSED")])
    ranges = [
        m for m in read.worksheet.merged_cells.ranges
        if m.min_col == COL_REMARKS and m.min_row == LEARNING_AREA_FIRST_ROW
    ]
    assert len(ranges) == 1
    assert ranges[0].max_row == LEARNING_AREA_FIRST_ROW


def test_exit_line_survives_a_save_and_reload_round_trip():
    """The dynamic unmerge/remerge must produce a file Excel (and openpyxl)
    can reopen cleanly, not just an in-memory object that looks right."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    strip_external_formulas(worksheet)
    _fill_learning_areas(
        worksheet, anchor_map(worksheet),
        [
            _row("Subject A", {1: D(88)}, D(88), "PASSED"),
            _row("Subject B", {1: D(75)}, D(75), "PASSED"),
        ],
        "Dropped as of 08/30/2026 due to Child labor, work",
    )
    saved = workbook_to_bytes(workbook)
    reloaded = openpyxl.load_workbook(io.BytesIO(saved))[SHEET_NAME]
    assert reloaded.cell(LEARNING_AREA_FIRST_ROW, COL_REMARKS).value == (
        "Dropped as of 08/30/2026 due to Child labor, work"
    )


# --- Identity fields -------------------------------------------------------


@pytest.mark.parametrize(
    "birthdate, reference, expected",
    [
        (date(2009, 1, 1), date(2026, 6, 8), 17),
        (date(2009, 6, 8), date(2026, 6, 8), 17),   # birthday on the day itself
        (date(2009, 6, 9), date(2026, 6, 8), 16),   # birthday the day after
        (date(2008, 12, 31), date(2026, 6, 8), 17),
    ],
)
def test_age_is_taken_at_the_school_year_start(birthdate, reference, expected):
    """Not "today" — a card reprinted years later must still show the age
    the learner was during that school year."""
    assert _age_on(birthdate, reference) == expected


def test_age_is_none_without_a_birthdate_or_reference():
    assert _age_on(None, date(2026, 6, 8)) is None
    assert _age_on(date(2009, 1, 1), None) is None


class _Level:
    def __init__(self, code, name):
        self.code, self.name = code, name


def test_grade_field_shows_the_number_only():
    """The form's field is labelled "Grade:", so it wants 11 — the stored
    code is "G11"."""
    assert _grade_level_number(_Level("G11", "Grade 11")) == "11"
    assert _grade_level_number(_Level("G12", "Grade 12")) == "12"
    assert _grade_level_number(None) == ""


# --- Template layout guard -------------------------------------------------


def test_template_still_matches_the_hardcoded_layout():
    """If the school swaps in a revised SF9 this fails loudly instead of
    silently writing grades into the wrong cells."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    assert SHEET_NAME in workbook.sheetnames
    worksheet = workbook[SHEET_NAME]

    assert worksheet["A18"].value == "Learning Areas"
    assert worksheet["H18"].value == "TERM"
    assert "Final" in str(worksheet["K18"].value)
    assert worksheet["L18"].value == "Remarks"
    assert [worksheet.cell(19, c).value for c in (8, 9, 10)] == [1, 2, 3]
    assert worksheet["A32"].value == "General Average"
    assert worksheet["O2"].value == "ATTENDANCE RECORD"
    assert worksheet["O4"].value == "No. of Class Days"
    assert worksheet["O5"].value == "No. of Days Present"
    assert worksheet["O6"].value == "No. of Days Absent"


def test_attendance_covers_a_june_to_april_school_year():
    assert ATTENDANCE_MONTHS == [6, 7, 8, 9, 10, 11, 12, 1, 2, 3, 4]
    assert len(ATTENDANCE_MONTHS) == 11
    assert [MONTH_ABBREVIATIONS[m] for m in ATTENDANCE_MONTHS[:3]] == ["Jun", "Jul", "Aug"]


def test_template_row_capacity_matches_the_constant():
    """MAX_LEARNING_AREAS is derived from the template's own bounds — the
    General Average row is what ends the block."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    assert worksheet["A32"].value == "General Average"
    assert MAX_LEARNING_AREAS == 32 - LEARNING_AREA_FIRST_ROW


# --- Term block-out flags -------------------------------------------------


def _flags(rows):
    """Column N as written — the template's conditional formatting decodes
    its three digits to decide which terms to block out."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    strip_external_formulas(worksheet)
    anchors = anchor_map(worksheet)
    _fill_learning_areas(worksheet, anchors, rows)

    def read(offset):
        row = LEARNING_AREA_FIRST_ROW + offset
        target = anchors.get((row, COL_TERM_OFFERED_FLAGS), (row, COL_TERM_OFFERED_FLAGS))
        return worksheet.cell(*target).value

    return read


def _one_term(name, term):
    return LearningAreaRow(
        name=name, term_grades={term: D(90)}, final_grade=D(90),
        remark="PASSED", offered_terms={term},
    )


@pytest.mark.parametrize(
    "offered, expected",
    [
        ({1, 2, 3}, 111),  # full year — nothing blocked out
        ({1}, 100),        # Term 1 only  -> Terms 2 and 3 blocked
        ({2}, 10),         # Term 2 only  -> Terms 1 and 3 blocked
        ({3}, 1),          # Term 3 only  -> Terms 1 and 2 blocked
        ({1, 2}, 110),
        ({2, 3}, 11),
    ],
)
def test_term_offered_flags_match_the_templates_encoding(offered, expected):
    """The rules read N as three per-term digits:
        H: INT($N/100)=0        -> block Term 1
        I: MOD(INT($N/10),10)=0 -> block Term 2
        J: MOD($N,10)=0         -> block Term 3
    """
    assert _term_flags(
        LearningAreaRow(
            name="Subject", term_grades={}, final_grade=None,
            remark=None, offered_terms=offered,
        )
    ) == expected


def test_a_one_term_elective_writes_the_right_flag():
    read = _flags([_one_term("Biology 1", 1)])
    assert read(0) == 100
    read = _flags([_one_term("Second-Term Elective", 2)])
    assert read(0) == 10
    read = _flags([_one_term("Third-Term Elective", 3)])
    assert read(0) == 1


def test_unused_rows_get_no_flag_so_they_stay_blank():
    """§35 wants spare rows visually blank. Every rule is guarded by
    $A<>"", so a nameless row never shades regardless — but leaving the
    flag empty keeps that unambiguous."""
    read = _flags([_one_term("Only", 1)])
    for offset in range(1, MAX_LEARNING_AREAS):
        assert read(offset) is None


def test_a_non_offered_term_is_blocked_out_and_an_offered_one_is_left_alone():
    """The block-out is now written as a direct fill as well as via the
    template's rules, because `app/xlsx_render.py` draws what a cell says
    rather than evaluating conditional formatting.

    The original bug this guards against was never "painting a fill" — it
    was *clearing* one: `PatternFill(fill_type=None)` serialises onto OOXML
    fill index 1, which is always gray125, so cells meant to be cleared
    came out grey. So the invariant is that an offered term is **left
    untouched**, never assigned a fill of any kind.
    """
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    strip_external_formulas(worksheet)
    assert len(worksheet.conditional_formatting) == 3
    _fill_learning_areas(worksheet, anchor_map(worksheet), [_one_term("Biology 1", 1)])

    row = LEARNING_AREA_FIRST_ROW
    offered = worksheet.cell(row=row, column=COL_TERM[1])
    blocked = worksheet.cell(row=row, column=COL_TERM[2])

    assert offered.fill.fill_type is None, "an offered term must keep the template's own styling"
    assert blocked.fill.fill_type == "solid"
    assert "595959" in str(blocked.fill.start_color.rgb), "must use the template's own block-out colour"


def test_the_block_out_never_produces_the_gray125_default():
    """The exact Phase 10 failure: any fill that serialises onto OOXML
    fill index 1 comes back as gray125 and paints cells that should have
    been left clear."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    strip_external_formulas(worksheet)
    _fill_learning_areas(worksheet, anchor_map(worksheet), [_one_term("Biology 1", 1)])

    saved = workbook_to_bytes(workbook)
    styles = zipfile.ZipFile(io.BytesIO(saved)).read("xl/styles.xml").decode()
    for row in range(LEARNING_AREA_FIRST_ROW, LEARNING_AREA_LAST_ROW + 1):
        for column in COL_TERM.values():
            cell = worksheet.cell(row=row, column=column)
            assert cell.fill.fill_type != "gray125"
    assert "gray125" in styles, "index 1 always exists; nothing may reference it"

    sheet = zipfile.ZipFile(io.BytesIO(saved)).read("xl/worksheets/sheet1.xml").decode("utf-8", "replace")
    assert sheet.count("<conditionalFormatting") == 3, "the template's own rules must survive"


# --- Certificate of Transfer ----------------------------------------------


@pytest.mark.parametrize(
    "grade, expected", [("11", "12"), ("12", "COLLEGE"), ("", ""), ("Grade", "")]
)
def test_next_level_after(grade, expected):
    """Grade 12 is the end of Senior High, so its completers are eligible
    for COLLEGE rather than a non-existent Grade 13."""
    assert next_level_after(grade) == expected


def test_print_setup_fits_the_card_on_a_single_landscape_page():
    """The template ships with no <pageSetup>, so the card would default
    to portrait at 100% and split across pages. Both fit dimensions are
    set to 1 so the whole card lands on one sheet, and fitToPage has to be
    set too or the fit fields silently do nothing."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    _apply_print_setup(worksheet)
    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.fitToWidth == 1
    assert worksheet.page_setup.fitToHeight == 1
    assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert "$AA$42" in worksheet.print_area


def test_margins_are_narrowed_to_limit_the_shrink():
    """The card is ~9.5in tall; landscape Letter inside the template's
    original 0.75in margins leaves only 7.0in, forcing a ~73% shrink to
    fit one page. 0.25in margins give ~8.0in, so it fits nearer 84%."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    assert worksheet.page_margins.top == 0.75  # as shipped
    _apply_print_setup(worksheet)
    margins = worksheet.page_margins
    assert margins.top == margins.bottom == 0.25
    assert margins.left == margins.right == 0.25


def test_card_is_centred_on_the_printed_page():
    """Fitting to one page scales by whichever dimension binds first —
    the height here — leaving horizontal slack that would otherwise park
    the card against the left margin."""
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    _apply_print_setup(worksheet)
    assert worksheet.print_options.horizontalCentered is True
    assert worksheet.print_options.verticalCentered is True
